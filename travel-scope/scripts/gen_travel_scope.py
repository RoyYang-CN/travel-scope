#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Travel Guide Builder - Generate Excel + HTML + Markdown travel guide files.

Usage:
    python gen_travel_scope.py --output-dir <dir> --map-platform <google|amap|baidu|all> --country <name>

The DATA section below must be edited by LLM before running.
Replace DESTINATIONS, HOTELS, RESTAURANTS, ATTRACTIONS, TRANSPORT, PRACTICAL, etc.
with real data collected for the target country.
"""

import argparse
import html as html_mod
import json
import math
import os
import sys
from urllib.parse import quote as url_quote

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ============================================================
# COORDINATE TRANSFORMATION (WGS84 <-> GCJ02 <-> BD09)
# ============================================================

_A = 6378245.0  # Semi-major axis
_EE = 0.00669342162296594323  # Eccentricity squared
_X_PI = math.pi * 3000.0 / 180.0


def _transform_lat(lng, lat):
    ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + 0.1 * lng * lat + 0.2 * math.sqrt(abs(lng))
    ret += (20.0 * math.sin(6.0 * lng * math.pi) + 20.0 * math.sin(2.0 * lng * math.pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lat * math.pi) + 40.0 * math.sin(lat / 3.0 * math.pi)) * 2.0 / 3.0
    ret += (160.0 * math.sin(lat / 12.0 * math.pi) + 320.0 * math.sin(lat * math.pi / 30.0)) * 2.0 / 3.0
    return ret


def _transform_lng(lng, lat):
    ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + 0.1 * lng * lat + 0.1 * math.sqrt(abs(lng))
    ret += (20.0 * math.sin(6.0 * lng * math.pi) + 20.0 * math.sin(2.0 * lng * math.pi)) * 2.0 / 3.0
    ret += (20.0 * math.sin(lng * math.pi) + 40.0 * math.sin(lng / 3.0 * math.pi)) * 2.0 / 3.0
    ret += (150.0 * math.sin(lng / 12.0 * math.pi) + 300.0 * math.sin(lng / 30.0 * math.pi)) * 2.0 / 3.0
    return ret


def wgs84_to_gcj02(lat, lng):
    """WGS84 -> GCJ02 (Mars coordinates). For China mainland only."""
    dlat = _transform_lat(lng - 105.0, lat - 35.0)
    dlng = _transform_lng(lng - 105.0, lat - 35.0)
    radlat = lat / 180.0 * math.pi
    magic = math.sin(radlat)
    magic = 1 - _EE * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((_A * (1 - _EE)) / (magic * sqrtmagic) * math.pi)
    dlng = (dlng * 180.0) / (_A / sqrtmagic * math.cos(radlat) * math.pi)
    mglat = lat + dlat
    mglng = lng + dlng
    return mglat, mglng


def gcj02_to_bd09(lat, lng):
    """GCJ02 -> BD09. For China mainland only."""
    z = math.sqrt(lng * lng + lat * lat) + 0.00002 * math.sin(lat * _X_PI)
    theta = math.atan2(lat, lng) + 0.000003 * math.cos(lng * _X_PI)
    bd_lng = z * math.cos(theta) + 0.0065
    bd_lat = z * math.sin(theta) + 0.006
    return bd_lat, bd_lng


def wgs84_to_bd09(lat, lng):
    """WGS84 -> BD09 (via GCJ02)."""
    gcj_lat, gcj_lng = wgs84_to_gcj02(lat, lng)
    return gcj02_to_bd09(gcj_lat, gcj_lng)


# ============================================================
# MAP URI GENERATORS
# ============================================================

def parse_coords(coord_str):
    """Parse 'lat,lng' string. Returns (lat, lng) floats."""
    parts = coord_str.split(",")
    if len(parts) != 2:
        return None, None
    try:
        return float(parts[0].strip()), float(parts[1].strip())
    except ValueError:
        return None, None


def gen_google_maps_uri(lat, lng, name=""):
    """Generate Google Maps search URI."""
    return f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"


def gen_amap_uri(lat, lng, name=""):
    """Generate Amap (Gaode) marker URI. Converts WGS84->GCJ02 for China."""
    # For overseas destinations, use raw coordinates
    # Heuristic: if lat > 55 or lat < 18 or lng < 70 or lng > 140, likely overseas -> use raw
    # Note: threshold changed from lat < 0 to lat < 18 to cover Philippines (~10°N)
    if lat > 55 or lat < 18 or lng < 70 or lng > 140:
        gcj_lat, gcj_lng = lat, lng
    else:
        gcj_lat, gcj_lng = wgs84_to_gcj02(lat, lng)
    encoded_name = url_quote(name) if name else ""
    return f"https://uri.amap.com/marker?position={gcj_lng},{gcj_lat}&name={encoded_name}"


def gen_baidu_maps_uri(lat, lng, name="", content=""):
    """Generate Baidu Maps marker URI. Converts WGS84->BD09 for China."""
    if lat > 55 or lat < 18 or lng < 70 or lng > 140:
        bd_lat, bd_lng = lat, lng
    else:
        bd_lat, bd_lng = wgs84_to_bd09(lat, lng)
    title = url_quote(name) if name else ""
    content_enc = url_quote(content) if content else ""
    return f"https://api.map.baidu.com/marker?location={bd_lat},{bd_lng}&title={title}&content={content_enc}&output=html&src=webapp.travelGuide"


def gen_map_uri(platform, lat, lng, name="", content=""):
    """Generate map URI for the specified platform."""
    if platform == "google":
        return gen_google_maps_uri(lat, lng, name)
    elif platform == "amap":
        return gen_amap_uri(lat, lng, name)
    elif platform == "baidu":
        return gen_baidu_maps_uri(lat, lng, name, content)
    return gen_google_maps_uri(lat, lng, name)


# ============================================================
# DATA SECTION
# LLM: Replace the sample data below with real data collected for the target country.
# ============================================================

COUNTRY = "菲律宾"  # Used in filenames and titles
COUNTRY_EN = "Philippines"
TRAVEL_MODE = ""  # Standard Q4: 自驾 / 非自驾
DOMESTIC_DESTINATION = False  # Set by domestic builders; overseas always audits required transport
OUTPUT_LANG = "cn"  # cn / en; the --language=both mode generates both artifacts
DEMO_MODE = False


def load_demo_fixture(path):
    """Load a sanitized, offline fixture without touching provider APIs."""
    global COUNTRY, COUNTRY_EN, TRAVEL_MODE, DOMESTIC_DESTINATION
    global DESTINATIONS, DESTINATION_COVERAGE, SOURCE_EVIDENCE, DYNAMIC_INFO
    global ROUTES, HOTELS, RESTAURANTS, TRANSPORT, ATTRACTIONS, DIVE_SITES
    global BUDGET, PRACTICAL, ITINERARY
    with open(path, "r", encoding="utf-8") as fixture_file:
        payload = json.load(fixture_file)
    if not isinstance(payload, dict):
        raise ValueError("Demo fixture must be a JSON object")
    COUNTRY = payload.get("country", COUNTRY)
    COUNTRY_EN = payload.get("country_en", COUNTRY_EN)
    TRAVEL_MODE = payload.get("travel_mode", "非自驾")
    DOMESTIC_DESTINATION = bool(payload.get("domestic_destination", False))
    for key in (
        "destinations", "destination_coverage", "source_evidence", "dynamic_info",
        "routes", "hotels", "restaurants", "transport", "attractions",
        "dive_sites", "budget", "practical", "itinerary",
    ):
        if key in payload:
            globals()[{
                "destinations": "DESTINATIONS", "destination_coverage": "DESTINATION_COVERAGE",
                "source_evidence": "SOURCE_EVIDENCE", "dynamic_info": "DYNAMIC_INFO",
                "routes": "ROUTES", "hotels": "HOTELS", "restaurants": "RESTAURANTS",
                "transport": "TRANSPORT", "attractions": "ATTRACTIONS", "dive_sites": "DIVE_SITES",
                "budget": "BUDGET", "practical": "PRACTICAL", "itinerary": "ITINERARY",
            }[key]] = payload[key]


def display_name(cn, en="", lang=None):
    """Return a traveller-facing name while retaining Chinese for wayfinding."""
    lang = lang or OUTPUT_LANG
    cn, en = str(cn or "").strip(), str(en or "").strip()
    if en and en.casefold() == cn.casefold():
        return cn or en
    if lang == "en" and en:
        return f"{en} ({cn})"
    return f"{cn} ({en})" if en and lang != "en" else (cn or en)


HEADER_EN = {
    "中文名": "Chinese Name", "英文名": "English Name", "区域": "Region",
    "核心体验": "Key Experiences", "建议天数": "Suggested Days", "最佳季节": "Best Season",
    "到达方式": "Arrival", "是否纳入环线": "Included in Route",
    "目的地": "Destination", "状态": "Status", "是否进路线": "In Route",
    "排除/待验证原因": "Exclusion / Verification Reason", "后续动作": "Next Action", "核验日期": "Verified Date",
    "方案名称": "Route Option", "天数": "Days", "路线": "Route", "特色": "Highlights",
    "适合人群": "Best For", "预算(元,不含机票)": "Budget (CNY, flights excluded)", "交通方式": "Transport",
    "酒店名称(中文)": "Hotel (Chinese)", "酒店名称(英文)": "Hotel (English)", "类型": "Type",
    "参考价(元/晚)": "Reference Price (CNY/night)", "评分(搜索补充)": "Rating (search supplement)",
    "评价数(旧字段)": "Review Count (legacy)", "经纬度(WGS84)": "Coordinates (WGS84)",
    "特色描述": "Highlights", "餐厅名称(中文)": "Restaurant (Chinese)", "餐厅名称(英文)": "Restaurant (English)",
    "特色菜品": "Signature Dishes", "人均(元)": "Average Spend (CNY)", "地址/位置": "Address / Location",
    "景点名称(中文)": "Attraction (Chinese)", "景点名称(英文)": "Attraction (English)",
    "门票/价格(元)": "Ticket / Price (CNY)", "推荐游玩时间": "Suggested Time", "描述": "Description",
    "运营公司": "Operator", "参考票价(元)": "Reference Fare (CNY)", "行程时间": "Travel Time",
    "班次频率": "Frequency", "出发地坐标": "Origin Coordinates", "目的地坐标": "Destination Coordinates",
    "备注": "Notes", "项目": "Item", "内容": "Details", "类别": "Category", "名称(中文)": "Name (Chinese)",
    "名称(英文)": "Name (English)", "经纬度": "Coordinates", "地图链接": "Map Link", "日序": "Day",
    "序号": "Sequence", "类型": "Type", "地点名称": "Place Name", "日期": "Date", "行程": "Plan",
    "住宿(元)": "Lodging (CNY)", "餐饮(元)": "Food (CNY)", "交通(元)": "Transport (CNY)",
    "活动/门票(元)": "Activities / Tickets (CNY)", "其他(元)": "Other (CNY)", "日合计(元)": "Daily Total (CNY)",
}

SHEET_EN = {
    "目的地总览": "Destinations", "目的地覆盖": "Destination Coverage", "来源证据": "Source Evidence",
    "动态信息": "Dynamic Information", "路线方案": "Route Options", "住宿推荐": "Accommodation",
    "美食推荐": "Food & Restaurants", "交通信息": "Transport", "景点活动": "Attractions & Activities",
    "潜水活动": "Diving", "预算汇总": "Budget Summary", "实用信息": "Practical Information",
    "坐标汇总": "Coordinates", "每日行程": "Daily Itinerary",
}


def localize_headers(headers):
    return [HEADER_EN.get(h, h) if OUTPUT_LANG == "en" else h for h in headers]


def localize_sheet_name(name):
    return SHEET_EN.get(name, name) if OUTPUT_LANG == "en" else name


def destination_display(name):
    for row in DESTINATIONS:
        if isinstance(row, dict) and row.get("name") == name:
            return display_name(row.get("name", ""), row.get("name_en", ""))
        if row and row[0] == name:
            return display_name(row[0], row[1] if len(row) > 1 else "")
    return name


def route_display(name):
    """Localize route labels while keeping the Chinese route key stable."""
    for row in ROUTES:
        if isinstance(row, dict) and row.get("name") == name:
            return display_name(row.get("name", ""), row.get("name_en", ""))
        if row and row[0] == name:
            return display_name(row[0], row[7] if len(row) > 7 else "")
    return name


def poi_display_html(item):
    """Render English first while retaining Chinese for wayfinding."""
    cn = html_mod.escape(str(item.get("cn", "")))
    en = html_mod.escape(str(item.get("en", "")))
    if OUTPUT_LANG == "en" and en:
        return f"{en} <span class=\"cn\">({cn})</span>"
    return f"{cn} <span class=\"en\">{en}</span>" if en else cn


def validate_english_dynamic_labels(destination_names):
    """Block incomplete EN output instead of silently publishing Chinese labels."""
    if OUTPUT_LANG != "en":
        return
    destination_en = {}
    for row in DESTINATIONS:
        if isinstance(row, dict):
            destination_en[row.get("name", "")] = row.get("name_en", "")
        elif row:
            destination_en[row[0]] = row[1] if len(row) > 1 else ""
    missing_destinations = [name for name in destination_names if not str(destination_en.get(name, "")).strip()]
    missing_routes = []
    for row in ROUTES:
        if not row:
            continue
        route_name = row.get("name", "") if isinstance(row, dict) else row[0]
        has_english = row.get("name_en", "") if isinstance(row, dict) else (row[7] if len(row) > 7 else "")
        if not str(has_english).strip():
            missing_routes.append(str(route_name))
    if missing_destinations or missing_routes:
        problems = []
        if missing_destinations:
            problems.append("destinations missing name_en: " + ", ".join(missing_destinations[:20]))
        if missing_routes:
            problems.append("routes missing name_en: " + ", ".join(missing_routes[:20]))
        raise ValueError("English output blocked: " + "; ".join(problems))


def country_display(lang=None):
    return COUNTRY_EN if (lang or OUTPUT_LANG) == "en" and COUNTRY_EN else COUNTRY

def effective_travel_mode():
    """Normalize missing/invalid Q4 conservatively to non-self-drive."""
    return TRAVEL_MODE if TRAVEL_MODE in {"自驾", "非自驾"} else "非自驾"

# Sheet 1: 目的地总览
# [中文名, 英文名, 区域, 核心体验, 建议天数, 最佳季节, 到达方式, 是否纳入环线]
DESTINATIONS = [
    ["宿务市", "Cebu City", "维萨亚斯", "历史文化/麦哲伦十字架/烤乳猪/购物", "2天", "11月-次年4月", "直飞(免签入境口岸)", "是"],
    # ... (LLM: add all destinations here)
]

# Sheet: 目的地覆盖
# [目的地, 状态, 是否进路线, 排除/待验证原因, 后续动作, 核验日期]
# 状态: 正式纳入 / 候选但排除 / 待验证
DESTINATION_COVERAGE = [
    ["宿务市", "正式纳入", "是", "与主线顺路", "收集酒店/餐厅/景点", ""],
    # ... (LLM: add every discovered major destination; do not silently omit candidates)
]

# Sheet: 来源证据
# [点位ID, 目的地, 点位名称, 来源URL, 来源类型, 核验日期, 支持事实, 是否独立来源, 冲突或疑问, 状态]
# 来源类型: 官方 / 地图平台 / 攻略 / 点评 / 新闻 / 交通运营方
# 支持事实: 存在 / 地址 / 营业 / 评分 / 价格 / 图片 / 坐标
SOURCE_EVIDENCE = [
    ["hotel-001", "宿务市", "示例酒店", "https://example.com/source", "地图平台", "", "存在/地址", "是", "", "待复核"],
    # ... (LLM: retain the full evidence ledger for every verified point)
]

# Sheet: 动态信息
# [信息ID, 目的地/路线, 项目, 内容, 最后核验日期, 动态等级, 出发前复核, 来源URL, 状态]
# 动态等级: 低 / 中 / 高；出发前复核: 是 / 否
DYNAMIC_INFO = [
    ["dynamic-001", "宿务市", "景点开放时间", "待核验", "", "高", "是", "https://example.com/source", "待复核"],
    # ... (LLM: add prices, schedules, hours, visa, safety and other changing facts)
]

# Sheet: 路线方案
# [方案名, 天数, 路线, 特色, 适合人群, 预算(元,不含机票), 交通方式, 英文方案名]
ROUTES = [
    ["方案A: 示例环线", "12天", "目的地A→目的地B→目的地C→目的地A", "特色描述", "适合人群", "5,000-7,000", "交通方式"],
    # ... (LLM: add all routes here)
]

# Sheet: 住宿推荐
# [目的地, 酒店名称(中文), 酒店名称(英文), 类型, 参考价(元/晚), 评分, 评价数, 经纬度, 图片URL1, 图片URL2, 特色描述]
# 评分格式(国内): 高德结构化评分；百度/搜索/AI字段独立保存
# 评分格式(海外): "Google:4.3"
# 标准深度: 每个目的地8家; 精简: 5家; 深度: 10家
HOTELS = [
    ["宿务市", "示例酒店", "Example Hotel", "4星酒店", "500", "Google:4.3", "1000", "10.2913,123.9494", "", "", "特色描述"],
    # ... (LLM: add real hotels here, 8 per destination for standard depth)
]
AI_RECOMMENDATIONS = {}

# Sheet: 美食推荐
# [目的地, 餐厅名称(中文), 餐厅名称(英文), 特色菜品, 人均(元), 评分, 评价数, 经纬度, 图片URL1, 图片URL2, 地址/位置]
RESTAURANTS = [
    ["宿务市", "示例餐厅", "Example Restaurant", "特色菜", "50-80", "Google:4.3", "500", "10.3000,123.9000", "", "", "地址"],
    # ... (LLM: add real restaurants here, 6 per destination for standard depth)
]

# Sheet: 交通信息
# [路线, 交通方式, 运营公司, 参考票价(元), 行程时间, 班次频率, 出发地坐标, 目的地坐标, 备注]
TRANSPORT = [
    ["城市A→城市B", "航班", "Example Airlines", "100-400", "1.5h", "每天5班", "10.0,123.0", "9.0,120.0", "备注"],
    # ... (LLM: add all transport routes here)
]

# Sheet: 景点活动
# [目的地, 景点名称(中文), 景点名称(英文), 类型, 门票/价格(元), 评分, 经纬度, 图片URL1, 图片URL2, 推荐游玩时间, 描述]
ATTRACTIONS = [
    ["宿务市", "示例景点", "Example Attraction", "历史遗迹", "免费", "Google:4.3", "10.2935,123.9010", "", "", "1小时", "描述"],
    # ... (LLM: add real attractions here, 6 per destination for standard depth)
]

# Sheet: 潜水活动 (如有)
# [目的地, 潜点名称(中文), 潜点名称(英文), 类型, 价格(元), 评分, 经纬度, 图片URL1, 图片URL2, 适合等级, 描述]
DIVE_SITES = [
    # ... (LLM: add dive sites if applicable)
]

# Sheet: 预算汇总 (示例路线)
# [天数, 日期, 行程, 住宿(元), 餐饮(元), 交通(元), 活动/门票(元), 其他(元), 日合计(元)]
BUDGET = [
    ["Day 1", "日期", "行程", 500, 80, 100, 0, 30, 710],
    # ... (LLM: add budget for the recommended route, numeric values as int, NOT string)
    ["合计", "N天", "全程", "=SUM(D2:D13)", "=SUM(E2:E13)", "=SUM(F2:F13)", "=SUM(G2:G13)", "=SUM(H2:H13)", "=SUM(I2:I13)"],
]

# Sheet: 实用信息
# [项目, 内容]
PRACTICAL = [
    ["签证政策", "签证信息"],
    ["汇率", "汇率信息"],
    ["SIM卡", "SIM卡信息"],
    ["最佳季节", "季节信息"],
    ["安全提示", "安全信息"],
    ["紧急联系", "紧急联系方式"],
    # ... (LLM: add all practical info here)
]

# Sheet: 每日行程 (用于地图路线连线, 仅推荐1-2条主路线)
# [方案名, 日序, 序号, 点位类型, 点位名称, 经纬度]
# 方案名: 对应ROUTES中的方案名
# 日序: "Day 1", "Day 2", ...
# 序号: 1, 2, 3, ... (每天内部的访问顺序)
# 点位类型: 景点/餐厅/酒店/交通
# 点位名称: 必须与HOTELS/RESTAURANTS/ATTRACTIONS中的名称一致(用于关联图片)
# 经纬度: WGS84格式 "纬度,经度"
ITINERARY = [
    # ["推荐路线", "Day 1", 1, "景点", "示例景点", "10.2935,123.9010"],
    # ["推荐路线", "Day 1", 2, "餐厅", "示例餐厅", "10.3000,123.9000"],
    # ["推荐路线", "Day 1", 3, "酒店", "示例酒店", "10.2913,123.9494"],
]

# Sheet: 每日行程 (用于地图路线连线, 仅推荐1-2条主路线)
# [方案名, 日序, 序号, 点位类型, 点位名称, 经纬度]
# 方案名: 对应ROUTES中的方案名
# 日序: "Day 1", "Day 2", ...
# 序号: 1, 2, 3, ... (每天内部的访问顺序)
# 点位类型: 景点/餐厅/酒店/交通
# 点位名称: 必须与HOTELS/RESTAURANTS/ATTRACTIONS中的名称一致(用于关联图片)
# 经纬度: WGS84格式 "纬度,经度"
ITINERARY = [
    # ["推荐路线", "Day 1", 1, "景点", "示例景点", "10.2935,123.9010"],
    # ["推荐路线", "Day 1", 2, "餐厅", "示例餐厅", "10.3000,123.9000"],
    # ["推荐路线", "Day 1", 3, "酒店", "示例酒店", "10.2913,123.9494"],
]


# ============================================================
# EXCEL GENERATION
# ============================================================

def generate_excel(output_path, include_research=True):
    """Generate a research or delivery workbook from the shared data section."""
    wb = openpyxl.Workbook()

    header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
    cell_font = Font(name='Microsoft YaHei', size=10)
    wrap_align = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    def write_sheet(ws, headers, data, col_widths=None):
        ws.title = localize_sheet_name(ws.title)
        headers = localize_headers(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = wrap_align
                cell.border = thin_border
        if col_widths:
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = 'A2'

    # Sheet 1: 目的地总览
    ws1 = wb.active
    ws1.title = '目的地总览'
    write_sheet(ws1,
        ['中文名', '英文名', '区域', '核心体验', '建议天数', '最佳季节', '到达方式', '是否纳入环线'],
        DESTINATIONS, [12, 18, 12, 30, 8, 14, 22, 12])

    if include_research:
        # Research Sheet 2: 目的地覆盖
        ws2 = wb.create_sheet('目的地覆盖')
        write_sheet(ws2,
            ['目的地', '状态', '是否进路线', '排除/待验证原因', '后续动作', '核验日期'],
            DESTINATION_COVERAGE, [20, 16, 12, 36, 28, 14])

        # Research Sheet 3: 来源证据
        ws3 = wb.create_sheet('来源证据')
        write_sheet(ws3,
            ['点位ID', '目的地', '点位名称', '来源URL', '来源类型', '核验日期', '支持事实', '是否独立来源', '冲突或疑问', '状态'],
            SOURCE_EVIDENCE, [16, 16, 24, 60, 18, 14, 26, 14, 30, 14])

        # Research Sheet 4: 动态信息
        ws4 = wb.create_sheet('动态信息')
        write_sheet(ws4,
            ['信息ID', '目的地/路线', '项目', '内容', '最后核验日期', '动态等级', '出发前复核', '来源URL', '状态'],
            DYNAMIC_INFO, [16, 20, 22, 42, 16, 12, 14, 60, 14])

    # Delivery/Research next sheet: 路线方案
    ws5 = wb.create_sheet('路线方案')
    write_sheet(ws5,
        ['方案名称', '天数', '路线', '特色', '适合人群', '预算(元,不含机票)', '交通方式'],
        ROUTES, [20, 6, 40, 28, 18, 16, 14])

    # Sheet 6: 住宿推荐
    ws6 = wb.create_sheet('住宿推荐')
    write_sheet(ws6,
        ['目的地', '酒店名称(中文)', '酒店名称(英文)', '类型', '参考价(元/晚)', '评分(搜索补充)', '评价数(旧字段)', '经纬度(WGS84)', '高德图片1', '高德图片2', '特色描述', '高德POI ID', '高德评分', '高德评价数', '百度UID', '百度匹配名称', '百度地址', '百度纬度', '百度经度', '百度评分', '百度评价数', '百度价格', '百度图片1', '百度图片2', '百度匹配状态', 'AI推荐分'],
        HOTELS, [12, 22, 30, 12, 14, 22, 12, 18, 30, 30, 35, 18, 12, 12, 22, 22, 35, 12, 12, 12, 12, 12, 30, 30, 16, 12])

    # Sheet 7: 美食推荐
    ws7 = wb.create_sheet('美食推荐')
    write_sheet(ws7,
        ['目的地', '餐厅名称(中文)', '餐厅名称(英文)', '特色菜品', '人均(元)', '评分(搜索补充)', '评价数(旧字段)', '经纬度(WGS84)', '高德图片1', '高德图片2', '地址/位置', '高德POI ID', '高德评分', '高德评价数', '百度UID', '百度匹配名称', '百度地址', '百度纬度', '百度经度', '百度评分', '百度评价数', '百度价格', '百度图片1', '百度图片2', '百度匹配状态', 'AI推荐分'],
        RESTAURANTS, [12, 20, 30, 22, 10, 22, 12, 18, 30, 30, 35, 18, 12, 12, 22, 22, 35, 12, 12, 12, 12, 12, 30, 30, 16, 12])

    # Sheet 8: 交通信息
    ws8 = wb.create_sheet('交通信息')
    write_sheet(ws8,
        ['路线', '交通方式', '运营公司', '参考票价(元)', '行程时间', '班次频率', '出发地坐标', '目的地坐标', '备注'],
        TRANSPORT, [22, 12, 28, 16, 14, 14, 16, 16, 28])

    # Sheet 9: 景点活动
    ws9 = wb.create_sheet('景点活动')
    write_sheet(ws9,
        ['目的地', '景点名称(中文)', '景点名称(英文)', '类型', '门票/价格(元)', '评分(搜索补充)', '经纬度(WGS84)', '高德图片1', '高德图片2', '推荐游玩时间', '描述', '高德POI ID', '高德评分', '高德评价数', '百度UID', '百度匹配名称', '百度地址', '百度纬度', '百度经度', '百度评分', '百度评价数', '百度价格', '百度图片1', '百度图片2', '百度匹配状态', 'AI推荐分'],
        ATTRACTIONS, [12, 22, 30, 12, 14, 22, 18, 30, 30, 12, 35, 18, 12, 12, 22, 22, 35, 12, 12, 12, 12, 12, 30, 30, 16, 12])

    # Sheet 10: 潜水活动
    ws10 = wb.create_sheet('潜水活动')
    write_sheet(ws10,
        ['目的地', '潜点名称(中文)', '潜点名称(英文)', '类型', '价格(元)', '评分', '经纬度', '图片1', '图片2', '适合等级', '描述'],
        DIVE_SITES, [12, 20, 28, 14, 14, 18, 16, 30, 30, 14, 30])

    # Sheet 11: 预算汇总
    ws11 = wb.create_sheet('预算汇总')
    write_sheet(ws11,
        ['天数', '日期', '行程', '住宿(元)', '餐饮(元)', '交通(元)', '活动/门票(元)', '其他(元)', '日合计(元)'],
        BUDGET, [8, 8, 22, 10, 10, 10, 14, 10, 10])

    # Sheet 12: 实用信息
    ws12 = wb.create_sheet('实用信息')
    write_sheet(ws12, ['项目', '内容'], PRACTICAL, [14, 80])

    # Sheet 13: 坐标汇总
    coords = []
    for h in HOTELS:
        coords.append(["住宿", h[0], h[1], h[2], h[7], gen_google_maps_uri(*parse_coords(h[7]))])
    for r in RESTAURANTS:
        coords.append(["美食", r[0], r[1], r[2], r[7], gen_google_maps_uri(*parse_coords(r[7]))])
    for a in ATTRACTIONS:
        coords.append(["景点", a[0], a[1], a[2], a[6], gen_google_maps_uri(*parse_coords(a[6]))])
    for t in TRANSPORT:
        dep = t[0].split('→')[0] if '→' in t[0] else t[0]
        arr = t[0].split('→')[1] if '→' in t[0] else ''
        coords.append(["交通-出发", dep, dep, "", t[6], gen_google_maps_uri(*parse_coords(t[6]))])
        if arr:
            coords.append(["交通-到达", arr, arr, "", t[7], gen_google_maps_uri(*parse_coords(t[7]))])

    ws13 = wb.create_sheet('坐标汇总')
    write_sheet(ws13,
        ['类别', '目的地', '名称(中文)', '名称(英文)', '经纬度', '地图链接'],
        coords, [10, 12, 22, 30, 18, 60])

    # Sheet 14: 每日行程
    ws14 = wb.create_sheet('每日行程')
    write_sheet(ws14,
        ['路线方案', '日序', '序号', '类型', '地点名称', '经纬度'],
        ITINERARY, [22, 8, 6, 8, 22, 18])

    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    print(f"Sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name}: {ws.max_row}r x {ws.max_column}c")
    return coords


# ============================================================
# HTML GENERATION (by destination, multi-map-platform)
# ============================================================

DEST_EMOJIS_DEFAULT = {
    "宿务市": "⛪", "墨宝": "🐟", "奥斯洛布": "🦈", "杜马盖蒂": "🐢",
    "锡基霍尔": "🌊", "薄荷岛": "🍫", "妈妈拍丝岛": "🦈", "长滩岛": "🏖",
    "公主港": "🕳", "爱妮岛": "🏝", "科隆": "⚓", "锡亚高岛": "🏄",
    "甘米银岛": "🌋", "大堡": "🦅", "维甘古城": "🏛", "马尼拉": "🏙",
}
CAT_EMOJIS = {"住宿": "🏨", "美食": "🍽", "景点": "📍", "潜水": "🤿", "交通": "🚌"}

MAP_PLATFORM_NAMES = {
    "google": "Google Maps",
    "amap": "高德地图",
    "baidu": "百度地图",
}
MAP_BTN_COLORS = {
    "google": "#1a73e8",
    "amap": "#008dfd",
    "baidu": "#2932e1",
}


def _is_overseas():
    """Check if destination is overseas based on first coordinate."""
    all_coords = []
    for h in HOTELS[:3]:
        all_coords.append(h[7])
    for a in ATTRACTIONS[:3]:
        all_coords.append(a[6])
    for c in all_coords:
        lat, lng = parse_coords(c)
        if lat is not None and lng is not None:
            if lat > 55 or lat < 18 or lng < 70 or lng > 140:
                return True
            return False
    return False


def _prepare_itinerary_json():
    """Build JSON data for map route view from ITINERARY array.
    Associates itinerary items with image URLs from HOTELS/RESTAURANTS/ATTRACTIONS."""
    if not ITINERARY:
        return json.dumps({})

    # Build name->images lookup
    img_lookup = {}
    for h in HOTELS:
        img_lookup[h[1]] = {"img1": h[8], "img2": h[9]}
    for r in RESTAURANTS:
        img_lookup[r[1]] = {"img1": r[8], "img2": r[9]}
    for a in ATTRACTIONS:
        img_lookup[a[1]] = {"img1": a[7], "img2": a[8]}

    routes_data = {}
    for item in ITINERARY:
        r_name, day, seq, ptype, pname, coords = item[0], item[1], item[2], item[3], item[4], item[5]
        if r_name not in routes_data:
            routes_data[r_name] = {}
        if day not in routes_data[r_name]:
            routes_data[r_name][day] = []
        imgs = img_lookup.get(pname, {"img1": "", "img2": ""})
        routes_data[r_name][day].append({
            "name": pname, "type": ptype, "coords": coords,
            "seq": seq, "img1": imgs["img1"], "img2": imgs["img2"]
        })

    return json.dumps(routes_data, ensure_ascii=False)


def generate_html(output_path, map_platforms, amap_js_key='', amap_security='', google_maps_key=''): 
    """Generate HTML with dual-view: card list + map route (dual-engine: AMap JS API for domestic / Google Maps JS API for overseas)."""
    guide_title = f"{country_display()} Travel Guide" if OUTPUT_LANG == "en" else f"{COUNTRY}旅行攻略"

    def valid_search_supplement(value):
        """Show only extracted search evidence; hide all-pending placeholders."""
        parts = [p.strip() for p in str(value or "").split("|")]
        hidden_statuses = ("待核验", "待验证", "未找到点位级证据", "找到点位证据但未提取到有效评分", "接口未提供")
        parts = [p for p in parts if p and not any(status in p for status in hidden_statuses)]
        return "｜".join(parts)

    # Organize items by destination -> category
    dest_items = {}
    for h in HOTELS:
        d = h[0]
        dest_items.setdefault(d, {"住宿": [], "美食": [], "景点": [], "潜水": [], "交通": []})
        dest_items[d]["住宿"].append({
            "cn": h[1], "en": h[2], "coords": h[7],
            "detail": f"{h[3]} | {h[4]}元/晚",
            "rating": h[5],
            "amap_rating": h[12] if len(h) > 12 else "",
            "amap_reviews": h[13] if len(h) > 13 else "",
            "baidu_rating": h[19] if len(h) > 19 else "",
            "baidu_reviews": h[20] if len(h) > 20 else "",
            "ai_score": h[25] if len(h) > 25 else "",
            "img1": h[8], "img2": h[9]
        })
    for r in RESTAURANTS:
        d = r[0]
        dest_items.setdefault(d, {"住宿": [], "美食": [], "景点": [], "潜水": [], "交通": []})
        dest_items[d]["美食"].append({
            "cn": r[1], "en": r[2], "coords": r[7],
            "detail": f"{r[3]} | 人均{r[4]}元",
            "rating": r[5],
            "amap_rating": r[12] if len(r) > 12 else "",
            "amap_reviews": r[13] if len(r) > 13 else "",
            "baidu_rating": r[19] if len(r) > 19 else "",
            "baidu_reviews": r[20] if len(r) > 20 else "",
            "ai_score": r[25] if len(r) > 25 else "",
            "img1": r[8], "img2": r[9]
        })
    for a in ATTRACTIONS:
        d = a[0]
        dest_items.setdefault(d, {"住宿": [], "美食": [], "景点": [], "潜水": [], "交通": []})
        dest_items[d]["景点"].append({
            "cn": a[1], "en": a[2], "coords": a[6],
            "detail": f"{a[3]} | {a[4]}元",
            "rating": a[5],
            "amap_rating": a[12] if len(a) > 12 else "",
            "amap_reviews": a[13] if len(a) > 13 else "",
            "baidu_rating": a[19] if len(a) > 19 else "",
            "baidu_reviews": a[20] if len(a) > 20 else "",
            "ai_score": a[25] if len(a) > 25 else "",
            "img1": a[7], "img2": a[8]
        })
    for dv in DIVE_SITES:
        d = dv[0]
        dest_items.setdefault(d, {"住宿": [], "美食": [], "景点": [], "潜水": [], "交通": []})
        dest_items[d]["潜水"].append({
            "cn": dv[1], "en": dv[2], "coords": dv[6],
            "detail": f"{dv[3]} | {dv[4]}元 | {dv[9]}",
            "rating": dv[5],
            "img1": dv[7], "img2": dv[8]
        })
    for t in TRANSPORT:
        route_str = t[0]  # e.g. "宿务→墨宝"
        dep = route_str.split('→')[0].strip() if '→' in route_str else route_str.strip()
        arr = route_str.split('→')[1].strip() if '→' in route_str else ''
        # Build transport detail: 方向 + 方式 + 公司 + 票价 + 耗时 + 班次
        transport_detail_parts = [t[1]]  # 交通方式
        if t[2]: transport_detail_parts.append(t[2])  # 运营公司
        if t[3]: transport_detail_parts.append(f"{t[3]}元")  # 参考票价
        if t[4]: transport_detail_parts.append(t[4])  # 行程时间
        if t[5]: transport_detail_parts.append(t[5])  # 班次频率
        transport_detail = ' | '.join(transport_detail_parts)
        transport_note = t[8] if len(t) > 8 and t[8] else ''  # 备注
        if transport_note:
            transport_detail += f"  [{transport_note}]"
        # Add to departure destination
        if dep in dest_items:
            dest_items[dep]["交通"].append({
                "cn": f"出发{dep} → 到达{arr}",
                "en": f"From {dep} to {arr}",
                "coords": t[6],
                "detail": transport_detail,
                "rating": ""
            })
        # Add to arrival destination
        if arr and arr in dest_items:
            dest_items[arr]["交通"].append({
                "cn": f"出发{dep} → 到达{arr}",
                "en": f"From {dep} to {arr}",
                "coords": t[6],  # use departure coords
                "detail": transport_detail,
                "rating": ""
            })

    # Generate map buttons HTML for an item
    def gen_map_buttons(coords_str, name_cn):
        lat, lng = parse_coords(coords_str)
        if lat is None:
            return ""
        buttons = []
        for platform in map_platforms:
            uri = gen_map_uri(platform, lat, lng, name_cn)
            btn_text = MAP_PLATFORM_NAMES[platform]
            btn_color = MAP_BTN_COLORS[platform]
            icon = {
                "google": '<svg viewBox="0 0 24 24" fill="currentColor" style="width:14px;height:14px"><path d="M12 2C8.13 2 5 5.13 5 9c0 5.25 7 13 7 13s7-7.75 7-13c0-3.87-3.13-7-7-7zm0 9.5c-1.38 0-2.5-1.12-2.5-2.5s1.12-2.5 2.5-2.5 2.5 1.12 2.5 2.5-1.12 2.5-2.5 2.5z"/></svg>',
                "amap": '<svg viewBox="0 0 24 24" fill="currentColor" style="width:14px;height:14px"><circle cx="12" cy="12" r="3"/></svg>',
                "baidu": '<svg viewBox="0 0 24 24" fill="currentColor" style="width:14px;height:14px"><rect x="6" y="6" width="12" height="12" rx="2"/></svg>',
            }.get(platform, '')
            buttons.append(
                f'<a class="btn-map btn-{platform}" href="{uri}" target="_blank" rel="noopener">{icon}{btn_text}</a>'
            )
        return ''.join(buttons)

    parts = []
    platform_names = " / ".join(MAP_PLATFORM_NAMES[p] for p in map_platforms)
    total_locations = len(HOTELS) + len(RESTAURANTS) + len(ATTRACTIONS) + len(DIVE_SITES)

    # Check map route availability
    overseas = _is_overseas()
    has_itinerary = bool(ITINERARY)
    map_route_enabled = (bool(amap_js_key) and not overseas or bool(google_maps_key) and overseas) and has_itinerary
    itinerary_json = _prepare_itinerary_json() if has_itinerary else "{}"
    all_route_names = json.dumps([r[0] for r in ROUTES], ensure_ascii=False)
    all_route_labels = json.dumps([route_display(r[0]) for r in ROUTES], ensure_ascii=False)
    _is_overseas_str = 'true' if overseas else 'false'

    # Build CSS for map buttons
    btn_css = ""
    for p in map_platforms:
        color = MAP_BTN_COLORS[p]
        btn_css += f".btn-{p} {{ background: {color}; }}"
        btn_css += f".btn-{p}:hover {{ background: {color}dd; }}\n"

    # AMap security config (if key provided)
    amap_security_script = ""
    if amap_js_key:
        sec = f'"{amap_security}"' if amap_security else '""'
        amap_security_script = f"""<script type="text/javascript">
window._AMapSecurityConfig = {{ securityJsCode: {sec} }};
</script>"""

    # Tab bar: show map tab only if enabled
    tab_bar_html = '<div class="tab-bar">'
    tab_bar_html += '<button class="tab-btn active" onclick="switchTab(\'list\')">📋 卡片列表</button>'
    if map_route_enabled:
        tab_bar_html += '<button class="tab-btn" onclick="switchTab(\'map\')">🗺️ 地图路线</button>'
    elif has_itinerary and not amap_js_key:
        tab_bar_html += '<button class="tab-btn tab-disabled" title="需要高德JS API Key">🗺️ 地图路线 (需Key)</button>'
    elif has_itinerary and overseas and not google_maps_key:
        tab_bar_html += '<button class="tab-btn tab-disabled" title="海外目的地需提供Google Maps API Key">🗺️ 地图路线 (海外)</button>'
    tab_bar_html += '</div>'

    parts.append(f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Travel-Scope | {guide_title}</title>
{amap_security_script}
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif; background: #f0f4f8; color: #333; line-height: 1.6; padding: 20px; }}
.container {{ max-width: 1200px; margin: 0 auto; }}
header {{ background: linear-gradient(135deg, #1a73e8, #0d47a1); color: white; padding: 30px; border-radius: 16px; margin-bottom: 24px; text-align: center; }}
header h1 {{ font-size: 24px; margin-bottom: 8px; }}
header p {{ font-size: 14px; opacity: 0.9; }}
header .stats {{ display: flex; justify-content: center; gap: 20px; margin-top: 16px; flex-wrap: wrap; }}
header .stat {{ background: rgba(255,255,255,0.15); padding: 8px 18px; border-radius: 8px; font-size: 14px; }}
header .stat strong {{ font-size: 20px; display: block; }}
#tab-list {{ display: block; }}
#tab-map {{ display: none; }}
.tab-bar {{ display: flex; gap: 0; margin-bottom: 20px; background: white; border-radius: 12px 12px 0 0; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
.tab-btn {{ flex: 1; padding: 14px 20px; border: none; background: #f5f5f5; color: #666; font-size: 15px; font-weight: 600; cursor: pointer; transition: all 0.2s; border-bottom: 3px solid transparent; }}
.tab-btn:hover {{ background: #e8f0fe; color: #1a73e8; }}
.tab-btn.active {{ background: white; color: #1a73e8; border-bottom-color: #1a73e8; }}
.tab-btn.tab-disabled {{ opacity: 0.5; cursor: not-allowed; }}
.toolbar {{ display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; align-items: center; }}
.btn-all {{ background: #1a73e8; color: white; border: none; padding: 10px 20px; border-radius: 8px; font-size: 14px; cursor: pointer; }}
.btn-all:hover {{ background: #1557b0; }}
.search-box {{ flex: 1; min-width: 200px; padding: 10px 16px; border: 2px solid #ddd; border-radius: 8px; font-size: 14px; outline: none; }}
.search-box:focus {{ border-color: #1a73e8; }}
.tip {{ background: #e8f5e9; border-left: 4px solid #4caf50; padding: 12px 16px; border-radius: 8px; margin-bottom: 20px; font-size: 14px; color: #2e7d32; }}
.dest-section {{ background: white; border-radius: 12px; margin-bottom: 20px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
.dest-header {{ padding: 16px 20px; cursor: pointer; display: flex; align-items: center; gap: 10px; font-size: 18px; font-weight: 600; border-bottom: 2px solid #f0f0f0; user-select: none; }}
.dest-header:hover {{ background: #f8f9fa; }}
.dest-header .emoji {{ font-size: 24px; }}
.dest-header .dest-count {{ background: #e8f0fe; color: #1a73e8; padding: 2px 12px; border-radius: 12px; font-size: 13px; font-weight: normal; }}
.dest-header .arrow {{ margin-left: auto; transition: transform 0.2s; color: #999; }}
.dest-section.collapsed .arrow {{ transform: rotate(-90deg); }}
.dest-section.collapsed .dest-body {{ display: none; }}
.cat-block {{ margin: 0; }}
.cat-header {{ padding: 10px 20px; background: #f8fbff; font-size: 14px; font-weight: 600; color: #1a73e8; border-bottom: 1px solid #eef2f7; display: flex; align-items: center; gap: 8px; cursor: pointer; user-select: none; }}
.cat-header:hover {{ background: #eef2f9; }}
.cat-header .cat-count {{ background: #e8f0fe; color: #1a73e8; padding: 1px 8px; border-radius: 10px; font-size: 12px; font-weight: normal; }}
.cat-header .cat-arrow {{ margin-left: auto; transition: transform 0.2s; color: #999; font-size: 12px; }}
.cat-block.collapsed .cat-arrow {{ transform: rotate(-90deg); }}
.cat-block.collapsed .cat-body {{ display: none; }}
.loc-card {{ display: flex; align-items: center; padding: 10px 20px; border-bottom: 1px solid #f5f5f5; gap: 12px; transition: background 0.15s; }}
.loc-card:hover {{ background: #f0f7ff; }}
.loc-card:last-child {{ border-bottom: none; }}
.loc-info {{ flex: 1; min-width: 0; }}
.loc-name {{ font-size: 14px; font-weight: 500; color: #202124; margin-bottom: 1px; }}
.loc-name .en {{ font-size: 12px; color: #80868b; font-weight: normal; margin-left: 6px; }}
.loc-detail {{ font-size: 12px; color: #5f6368; margin-bottom: 1px; }}
.loc-coords {{ font-size: 11px; color: #80868b; font-family: 'Roboto Mono', monospace; }}
.loc-rating {{ display: inline-block; background: #fef7e0; color: #f9ab00; padding: 1px 6px; border-radius: 4px; font-size: 11px; font-weight: 600; margin-left: 6px; }}
.loc-ai-rating {{ display: inline-block; background: #eef7ff; color: #2471a3; padding: 1px 6px; border-radius: 4px; font-size: 11px; font-weight: 600; margin-left: 6px; }}
.btn-group {{ display: flex; gap: 6px; flex-shrink: 0; }}
.btn-map {{ color: white; border: none; padding: 7px 12px; border-radius: 6px; font-size: 12px; cursor: pointer; text-decoration: none; display: inline-flex; align-items: center; gap: 4px; white-space: nowrap; transition: all 0.2s; }}
.btn-map:hover {{ transform: translateY(-1px); box-shadow: 0 2px 6px rgba(0,0,0,0.2); }}
{btn_css}
.loc-thumbs {{ display: flex; gap: 6px; flex-shrink: 0; }}
.loc-thumb-wrap {{ position: relative; width: 120px; height: 80px; border-radius: 8px; overflow: hidden; flex-shrink: 0; background: #f0f0f0; }}
.loc-thumb {{ width: 100%; height: 100%; object-fit: cover; cursor: pointer; transition: transform 0.2s; position: relative; z-index: 1; }}
.loc-thumb:hover {{ transform: scale(1.08); }}
.loc-thumb-fallback {{ position: absolute; inset: 0; display: flex; align-items: center; justify-content: center; background: #f0f0f0; font-size: 24px; color: #bbb; }}
.lightbox {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.85); z-index: 9999; justify-content: center; align-items: center; cursor: pointer; }}
.lightbox.active {{ display: flex; }}
.lightbox img {{ max-width: 90vw; max-height: 90vh; object-fit: contain; border-radius: 8px; }}
footer {{ text-align: center; padding: 20px; color: #999; font-size: 13px; }}
/* Map route view */
.map-container {{ display: flex; gap: 0; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.06); height: 700px; }}
.itinerary-panel {{ width: 300px; flex-shrink: 0; border-right: 1px solid #e0e0e0; overflow-y: auto; background: #fafbff; }}
.itinerary-panel h3 {{ padding: 16px; font-size: 16px; color: #1a73e8; border-bottom: 2px solid #e8f0fe; }}
.route-selector {{ padding: 12px 16px; border-bottom: 1px solid #eef2f7; }}
.route-selector select {{ width: 100%; padding: 8px 12px; border: 2px solid #ddd; border-radius: 8px; font-size: 14px; outline: none; }}
.route-selector select:focus {{ border-color: #1a73e8; }}
.day-group {{ margin: 0; }}
.day-header {{ padding: 10px 16px; font-size: 14px; font-weight: 600; color: #333; background: #f0f4f8; border-bottom: 1px solid #e8eef5; display: flex; align-items: center; gap: 8px; }}
.day-header .day-dot {{ width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }}
.itin-item {{ padding: 8px 16px 8px 32px; font-size: 13px; color: #5f6368; border-bottom: 1px solid #f5f5f5; cursor: pointer; display: flex; align-items: center; gap: 6px; transition: background 0.15s; }}
.itin-item:hover {{ background: #e8f0fe; }}
.itin-item.active {{ background: #e8f0fe; color: #1a73e8; font-weight: 600; }}
.itin-item .type-icon {{ font-size: 14px; flex-shrink: 0; }}
.itin-item .seq-num {{ display: inline-block; min-width: 20px; height: 20px; line-height: 20px; text-align: center; border-radius: 50%; background: #1a73e8; color: white; font-size: 11px; font-weight: 700; flex-shrink: 0; }}
.map-canvas-wrapper {{ flex: 1; position: relative; }}
#amap-container {{ width: 100%; height: 100%; }}
.map-legend {{ position: absolute; bottom: 16px; right: 16px; background: white; padding: 10px 16px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.15); font-size: 12px; z-index: 1000; }}
.map-legend .legend-item {{ display: flex; align-items: center; gap: 6px; margin: 2px 0; }}
.map-legend .legend-dot {{ width: 12px; height: 12px; border-radius: 50%; flex-shrink: 0; }}
.map-legend .legend-line {{ width: 16px; height: 3px; flex-shrink: 0; }}
.map-info-window {{ min-width: 200px; }}
.map-info-window h4 {{ font-size: 15px; color: #202124; margin-bottom: 6px; }}
.map-info-window .info-detail {{ font-size: 12px; color: #5f6368; margin-bottom: 8px; }}
.map-info-window .info-thumbs {{ display: flex; gap: 4px; margin-bottom: 8px; }}
.map-info-window .info-thumb {{ width: 80px; height: 54px; border-radius: 6px; object-fit: cover; cursor: pointer; }}
.map-info-window .info-thumb-placeholder {{ width: 80px; height: 54px; border-radius: 6px; background: #f0f0f0; display: flex; align-items: center; justify-content: center; font-size: 18px; color: #bbb; }}
.map-info-window .info-nav {{ display: flex; gap: 4px; }}
.map-info-window .info-nav a {{ display: inline-block; padding: 4px 10px; border-radius: 4px; font-size: 12px; color: white; text-decoration: none; }}
</style>
</head>
<body>
<div class="container">
<header>
<h1>Travel-Scope | {guide_title}</h1>
<p>互动探索 · {platform_names}一键导航 · 按目的地分类{' · 含地图路线' if map_route_enabled else ''}</p>
<div class="stats">
<div class="stat"><strong>{len(HOTELS)}</strong>住宿</div>
<div class="stat"><strong>{len(RESTAURANTS)}</strong>美食</div>
<div class="stat"><strong>{len(ATTRACTIONS)}</strong>景点</div>
<div class="stat"><strong>{len(DIVE_SITES)}</strong>潜水</div>
<div class="stat"><strong>{len(dest_items)}</strong>目的地</div>
</div>
</header>
{tab_bar_html}
<div id="tab-list">
<div class="toolbar">
<button class="btn-all" onclick="toggleAll()">展开/收起全部</button>
<input class="search-box" type="text" placeholder="搜索目的地/酒店/餐厅/景点..." oninput="filterCards(this.value)">
</div>
<div class="tip">点击地点右侧 <strong>{platform_names}</strong> 按钮可一键导航；点击目的地标题可展开/收起</div>
""")

    # Build destination sections. Include data-backed destinations even when the
    # overview table is a higher-level country summary and does not enumerate
    # every collected area.
    destination_names = list(dict.fromkeys([d[0] for d in DESTINATIONS] + list(dest_items.keys())))
    validate_english_dynamic_labels(destination_names)
    for dest_name in destination_names:
        items = dest_items.get(dest_name, {})
        emoji = DEST_EMOJIS_DEFAULT.get(dest_name, "📍")
        total = sum(len(v) for v in items.values())
        if total == 0:
            continue

        parts.append(f'\n<div class="dest-section" data-dest="{html_mod.escape(dest_name)}">')
        parts.append(f'<div class="dest-header" onclick="toggleSection(this)"><span class="emoji">{emoji}</span> {html_mod.escape(destination_display(dest_name))} <span class="dest-count">{total}个地点</span> <span class="arrow">▼</span></div>')
        parts.append('<div class="dest-body">')

        for cat in ["住宿", "美食", "景点", "潜水", "交通"]:
            cat_list = items.get(cat, [])
            if not cat_list:
                continue
            ce = CAT_EMOJIS.get(cat, "📍")
            parts.append(f'<div class="cat-block" data-cat="{cat}"><div class="cat-header" onclick="toggleCat(this)">{ce} {cat} <span class="cat-count">{len(cat_list)}</span><span class="cat-arrow">▼</span></div><div class="cat-body">')

            for item in cat_list:
                buttons = gen_map_buttons(item["coords"], item["cn"])
                search_supplement = valid_search_supplement(item.get("rating"))
                if overseas and search_supplement:
                    parts_rating = search_supplement.split("|", 1)
                    google_rating = parts_rating[0].replace("Google:", "").strip()
                    google_reviews = parts_rating[1].replace("条", "").strip() if len(parts_rating) > 1 else ""
                    google_suffix = f"｜评价数：{html_mod.escape(google_reviews)}" if google_reviews else ""
                    rating_html = f'<span class="loc-rating">Google评分：{html_mod.escape(google_rating)}{google_suffix}</span>'
                else:
                    rating_html = f'<span class="loc-rating">搜索补充：{html_mod.escape(search_supplement)}</span>' if search_supplement else ""
                amap_review_suffix = f'｜评价数：{html_mod.escape(item["amap_reviews"])}' if item.get("amap_reviews") and item.get("amap_reviews") not in {"待核验", "待验证", "接口未提供"} else ""
                amap_html = f'<span class="loc-rating">高德评分：{html_mod.escape(item["amap_rating"])}{amap_review_suffix}</span>' if item.get("amap_rating") else ""
                baidu_html = ""
                ai_html = f'<span class="loc-ai-rating">AI推荐分 {html_mod.escape(item["ai_score"])}（AI提供，仅供参考；非平台评分）</span>' if item.get("ai_score") else ""
                search_text = html_mod.escape(dest_name + " " + item["cn"] + " " + item.get("en", "").lower())
                thumbs_html = ""
                if cat != "交通":
                    thumbs_inner = ""
                    for img_key in ["img1", "img2"]:
                        img_url = item.get(img_key, "")
                        if img_url:
                            thumbs_inner += f'<div class="loc-thumb-wrap"><img class="loc-thumb" src="{html_mod.escape(img_url)}" alt="{html_mod.escape(item["cn"])}" onclick="openLightbox(this.src)" onerror="this.hidden=true"><div class="loc-thumb-fallback">📷</div></div>'
                        else:
                            thumbs_inner += '<div class="loc-thumb-wrap"><div class="loc-thumb-fallback">📷</div></div>'
                    thumbs_html = f'<div class="loc-thumbs">{thumbs_inner}</div>'
                parts.append(f"""<div class="loc-card" data-search="{search_text}">
{thumbs_html}<div class="loc-info">
<div class="loc-name">{poi_display_html(item)}{rating_html}{amap_html}{baidu_html}{ai_html}</div>
<div class="loc-detail">{html_mod.escape(item['detail'])}</div>
<div class="loc-coords">{item['coords']}</div>
</div>
<div class="btn-group">{buttons}</div>
</div>""")

            parts.append('</div></div>')  # close cat-body and cat-block

        parts.append('</div></div>')

    # Close tab-list, start tab-map
    map_view_html = f"""
</div><!-- /tab-list -->

<div id="tab-map">
<div class="map-container">
  <div class="itinerary-panel" id="itin-panel">
    <h3>🗺️ 行程路线</h3>
    <div class="route-selector">
      <select id="route-select" onchange="switchRoute()"></select>
    </div>
    <div id="itin-list"></div>
  </div>
  <div class="map-canvas-wrapper">
    <div id="amap-container"></div>
    <div class="map-legend" id="map-legend">
      <div class="legend-item"><span class="legend-dot" style="background:#4CAF50"></span> 景点</div>
      <div class="legend-item"><span class="legend-dot" style="background:#FF9800"></span> 餐厅</div>
      <div class="legend-item"><span class="legend-dot" style="background:#2196F3"></span> 酒店</div>
      <div class="legend-item"><span class="legend-dot" style="background:#9E9E9E"></span> 交通</div>
    </div>
  </div>
</div>
</div><!-- /tab-map -->
"""

    parts.append(map_view_html)

    parts.append(f"""
<footer>
<p>{guide_title} · 地图标记页 · 数据仅供参考</p>
<p>导航平台: {platform_names}</p>
</footer>
</div>
<div class="lightbox" onclick="closeLightbox()"><img id="lightbox-img" src=""></div>
<script>
// === Card list view JS ===
function openLightbox(src) {{
  document.getElementById('lightbox-img').src = src;
  document.querySelector('.lightbox').classList.add('active');
}}
function closeLightbox() {{
  document.querySelector('.lightbox').classList.remove('active');
}}
function toggleSection(el) {{ el.parentElement.classList.toggle('collapsed'); }}
function toggleCat(el) {{ el.parentElement.classList.toggle('collapsed'); }}
function toggleAll() {{
  const sections = document.querySelectorAll('.dest-section');
  const allCollapsed = [...sections].every(s => s.classList.contains('collapsed'));
  sections.forEach(s => {{ if (allCollapsed) {{ s.classList.remove('collapsed'); s.querySelectorAll('.cat-block').forEach(c => c.classList.remove('collapsed')); }} else {{ s.classList.add('collapsed'); }} }});
}}
function filterCards(query) {{
  const q = query.toLowerCase();
  document.querySelectorAll('.dest-section').forEach(sec => {{
    let hasVisible = false;
    sec.querySelectorAll('.loc-card').forEach(card => {{
      const match = card.dataset.search.toLowerCase().includes(q);
      card.style.display = match ? '' : 'none';
      if (match) hasVisible = true;
    }});
    sec.style.display = hasVisible ? '' : 'none';
    if (q && hasVisible) sec.classList.remove('collapsed');
  }});
}}

// === Tab switching ===
function switchTab(tab) {{
  document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(btn => {{
    if ((tab === 'list' && btn.textContent.includes('卡片')) ||
        (tab === 'map' && btn.textContent.includes('地图'))) {{
      btn.classList.add('active');
    }}
  }});
  document.getElementById('tab-list').style.display = tab === 'list' ? 'block' : 'none';
  document.getElementById('tab-map').style.display = tab === 'map' ? 'block' : 'none';
  if (tab === 'map' && !window._mapInit) {{
    initMap();
    window._mapInit = true;
  }}
}}

    // === Map route view (dual engine: AMap domestic / Google Maps overseas) ===
var itineraryData = {itinerary_json};
var allRouteNames = {all_route_names};
var allRouteLabels = {all_route_labels};
var mapInstance = null;
var markersLayer = null;
var polylinesLayer = null;
var currentRoute = null;
var currentMarkers = [];
var currentPolylines = [];
var currentOverlays = [];

var DAY_COLORS = ['#E53935', '#1E88E5', '#43A047', '#8E24AA', '#FB8C00', '#00ACC1', '#3949AB', '#D81B60'];
var TYPE_COLORS = {{'景点': '#4CAF50', '餐厅': '#FF9800', '酒店': '#2196F3', '交通': '#9E9E9E'}};
var TYPE_ICONS = {{'景点': '📍', '餐厅': '🍽️', '酒店': '🏨', '交通': '🚌'}};

var IS_OVERSEAS = {_is_overseas_str};
var GOOGLE_MAPS_KEY = "{google_maps_key}";
var MAP_ENGINE = IS_OVERSEAS && GOOGLE_MAPS_KEY ? "google" : "amap";

function initMap() {{
  if (MAP_ENGINE === "google") {{
    initGoogleMap();
  }} else {{
    initAMap();
  }}
}}

// === AMap (domestic China) ===
function initAMap() {{
  var script = document.createElement('script');
  script.src = 'https://webapi.amap.com/maps?v=2.0&key={amap_js_key}';
  script.onload = function() {{
    mapInstance = new AMap.Map('amap-container', {{
      zoom: 6,
      mapStyle: 'amap://styles/normal',
      viewMode: '2D'
    }});
    mapInstance._engine = 'amap';
    // Populate route selector with ALL routes
    var select = document.getElementById('route-select');
    allRouteNames.forEach(function(name, i) {{
      var opt = document.createElement('option');
      opt.value = name;
       opt.textContent = allRouteLabels[i] || name;
      select.appendChild(opt);
    }});
    if (allRouteNames.length > 0) {{
      currentRoute = allRouteNames[0];
      renderRoute(currentRoute);
    }}
  }};
  document.head.appendChild(script);
}}

// === Google Maps (overseas) ===
function initGoogleMap() {{
  var script = document.createElement('script');
  script.src = 'https://maps.googleapis.com/maps/api/js?key=' + GOOGLE_MAPS_KEY + '&callback=_gmapReady';
  window._gmapReady = function() {{
    mapInstance = new google.maps.Map(document.getElementById('amap-container'), {{
      zoom: 6,
      mapTypeId: google.maps.MapTypeId.ROADMAP,
      center: {{lat: 0, lng: 0}}
    }});
    mapInstance._engine = 'google';
    // Populate route selector
    var select = document.getElementById('route-select');
    allRouteNames.forEach(function(name, i) {{
      var opt = document.createElement('option');
      opt.value = name;
       opt.textContent = allRouteLabels[i] || name;
      select.appendChild(opt);
    }});
    if (allRouteNames.length > 0) {{
      currentRoute = allRouteNames[0];
      renderRoute(currentRoute);
    }}
  }};
  document.head.appendChild(script);
}}

function switchRoute() {{
  currentRoute = document.getElementById('route-select').value;
  renderRoute(currentRoute);
}}

function renderRoute(routeName) {{
  // Clear previous
  if (mapInstance._engine === 'google') {{
    currentMarkers.forEach(function(m) {{ m.setMap(null); }});
    currentPolylines.forEach(function(p) {{ p.setMap(null); }});
    currentOverlays.forEach(function(o) {{ o.setMap(null); }});
  }} else {{
    currentMarkers.forEach(function(m) {{ mapInstance.remove(m); }});
    currentPolylines.forEach(function(p) {{ mapInstance.remove(p); }});
  }}
  currentMarkers = [];
  currentPolylines = [];
  currentOverlays = [];

  var routeData = itineraryData[routeName];
  if (!routeData) {{
    document.getElementById('itin-list').innerHTML = '<div style="padding:24px 16px;color:#999;text-align:center;font-size:14px;">📋 该路线暂无详细行程编排<br><span style="font-size:12px;color:#bbb;">仅推荐路线提供每日行程地图展示</span></div>';
    return;
  }}

  var allPoints = [];
  var days = Object.keys(routeData).sort(function(a, b) {{
    var na = parseInt(a.replace(/\\D/g, ''), 10) || 0;
    var nb = parseInt(b.replace(/\\D/g, ''), 10) || 0;
    return na - nb;
  }});

  // Build itinerary panel
  var panelHtml = '';
  var globalIdx = 0;
  days.forEach(function(day, dayIdx) {{
    var color = DAY_COLORS[dayIdx % DAY_COLORS.length];
    var items = routeData[day];
    panelHtml += '<div class="day-group">';
    panelHtml += '<div class="day-header"><span class="day-dot" style="background:' + color + '"></span>' + day + '</div>';
    items.forEach(function(item, itemIdx) {{
      var letter = String.fromCharCode(65 + itemIdx); // A, B, C...
      item._label = day + '-' + letter; // e.g. "Day 1-A"
      item._letter = letter;
      item._dayIdx = dayIdx;
      item._dayFirst = (itemIdx === 0);
      allPoints.push(item);
      var typeIcon = TYPE_ICONS[item.type] || '📍';
      panelHtml += '<div class="itin-item" onclick="flyTo(' + globalIdx + ')">';
      panelHtml += '<span class="seq-num" style="background:' + color + '">' + letter + '</span>';
      panelHtml += '<span class="type-icon">' + typeIcon + '</span>';
      panelHtml += '<span>' + item.name + '</span>';
      panelHtml += '</div>';
      globalIdx++;
    }});
    panelHtml += '</div>';
  }});
  document.getElementById('itin-list').innerHTML = panelHtml;

  // Update legend with actual days
  var legendEl = document.getElementById('map-legend');
  var typeLegendHtml = '<div class="legend-item"><span class="legend-dot" style="background:#4CAF50"></span> 景点</div>' +
    '<div class="legend-item"><span class="legend-dot" style="background:#FF9800"></span> 餐厅</div>' +
    '<div class="legend-item"><span class="legend-dot" style="background:#2196F3"></span> 酒店</div>' +
    '<div class="legend-item"><span class="legend-dot" style="background:#9E9E9E"></span> 交通</div>';
  var dayLegendHtml = '';
  days.forEach(function(day, dayIdx) {{
    var color = DAY_COLORS[dayIdx % DAY_COLORS.length];
    dayLegendHtml += '<div class="legend-item"><span class="legend-line" style="background:' + color + '"></span> ' + day + '</div>';
  }});
  if (legendEl) legendEl.innerHTML = typeLegendHtml + dayLegendHtml;

  // Add markers based on engine type
  if (mapInstance._engine === 'google') {{
    renderRouteGoogle(allPoints, routeData, days);
  }} else {{
    renderRouteAMap(allPoints, routeData, days);
  }}
}}

function renderRouteAMap(allPoints, routeData, days) {{
  allPoints.forEach(function(item, idx) {{
    var parts = item.coords.split(',');
    var lat = parseFloat(parts[0]);
    var lng = parseFloat(parts[1]);
    var gcj = wgs84ToGcj02(lat, lng);
    var marker = new AMap.Marker({{
      position: [gcj.lng, gcj.lat],
      title: item.name,
      label: {{
        content: '<div style="background:' + (DAY_COLORS[item._dayIdx % DAY_COLORS.length]) + ';color:white;width:22px;height:22px;line-height:22px;text-align:center;border-radius:50%;font-size:12px;font-weight:700;">' + (item._letter || item.seq) + '</div>',
        direction: 'top'
      }}
    }});
    // Day label marker for first point of each day
    var isFirstOfDay = item._dayFirst;
    if (isFirstOfDay) {{
      var dayColor = DAY_COLORS[item._dayIdx % DAY_COLORS.length];
      var dayMarker = new AMap.Marker({{
        position: [gcj.lng, gcj.lat],
        content: '<div style="position:relative;top:-32px;left:16px;background:' + dayColor + ';color:white;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;white-space:nowrap;box-shadow:0 1px 4px rgba(0,0,0,0.3);pointer-events:none;">' + ('Day ' + (item._dayIdx + 1)) + '</div>',
        offset: new AMap.Pixel(0, 0)
      }});
      currentMarkers.push(dayMarker);
      mapInstance.add(dayMarker);
    }}

    var thumbHtml = '';
    if (item.img1) {{
      thumbHtml += '<img class="info-thumb" src="' + item.img1 + '" onerror="this.style.display=\\'none\\'" onclick="openLightbox(this.src)">';
    }} else {{
      thumbHtml += '<div class="info-thumb-placeholder">📷</div>';
    }}
    if (item.img2) {{
      thumbHtml += '<img class="info-thumb" src="' + item.img2 + '" onerror="this.style.display=\\'none\\'" onclick="openLightbox(this.src)">';
    }}
    var navLinks = '';
    var navPlatforms = {json.dumps(map_platforms)};
    navPlatforms.forEach(function(p) {{
      var uri = genNavUri(p, lat, lng, item.name);
      var names = {{'google':'Google','amap':'高德','baidu':'百度'}};
      var colors = {{'google':'#1a73e8','amap':'#008dfd','baidu':'#2932e1'}};
      navLinks += '<a href="' + uri + '" target="_blank" style="background:' + colors[p] + '">' + names[p] + '</a>';
    }});

    var infoContent = '<div class="map-info-window">' +
      '<h4>' + typeIconHtml(item) + ' ' + item.name + '</h4>' +
      '<div class="info-detail">类型: ' + item.type + '</div>' +
      '<div class="info-thumbs">' + thumbHtml + '</div>' +
      '<div class="info-nav">' + navLinks + '</div>' +
      '</div>';

    marker.infoWindow = new AMap.InfoWindow({{
      content: infoContent,
      offset: new AMap.Pixel(0, -30)
    }});
    marker.on('click', function(e) {{
      marker.infoWindow.open(mapInstance, marker.getPosition());
    }});
    currentMarkers.push(marker);
    mapInstance.add(marker);
  }});

  // Draw polylines (AMap)
  days.forEach(function(day, dayIdx) {{
    var color = DAY_COLORS[dayIdx % DAY_COLORS.length];
    var items = routeData[day];
    var path = [];
    items.forEach(function(item) {{
      var parts = item.coords.split(',');
      var lat = parseFloat(parts[0]);
      var lng = parseFloat(parts[1]);
      var gcj = wgs84ToGcj02(lat, lng);
      path.push([gcj.lng, gcj.lat]);
    }});
    if (path.length >= 2) {{
      var polyline = new AMap.Polyline({{
        path: path,
        strokeColor: color,
        strokeWeight: 3,
        strokeOpacity: 0.8,
        strokeStyle: 'solid',
        showDir: true
      }});
      currentPolylines.push(polyline);
      mapInstance.add(polyline);
    }}
  }});

  // Dashed lines between days (AMap)
  for (var i = 0; i < days.length - 1; i++) {{
    var dayItems1 = routeData[days[i]];
    var dayItems2 = routeData[days[i + 1]];
    var lastItem = dayItems1[dayItems1.length - 1];
    var firstItem = dayItems2[0];
    var p1 = lastItem.coords.split(',');
    var p2 = firstItem.coords.split(',');
    var gcj1 = wgs84ToGcj02(parseFloat(p1[0]), parseFloat(p1[1]));
    var gcj2 = wgs84ToGcj02(parseFloat(p2[0]), parseFloat(p2[1]));
    var dashLine = new AMap.Polyline({{
      path: [[gcj1.lng, gcj1.lat], [gcj2.lng, gcj2.lat]],
      strokeColor: '#999',
      strokeWeight: 2,
      strokeOpacity: 0.6,
      strokeStyle: 'dashed'
    }});
    currentPolylines.push(dashLine);
    mapInstance.add(dashLine);
  }}

  if (currentMarkers.length > 0) {{
    mapInstance.setFitView(currentMarkers.concat(currentPolylines), false, [60, 60, 60, 60]);
  }}
}}

function renderRouteGoogle(allPoints, routeData, days) {{
  var bounds = new google.maps.LatLngBounds();

  allPoints.forEach(function(item, idx) {{
    var parts = item.coords.split(',');
    var lat = parseFloat(parts[0]);
    var lng = parseFloat(parts[1]);

    var marker = new google.maps.Marker({{
      position: {{lat: lat, lng: lng}},
      title: item.name,
      icon: {{
        path: google.maps.SymbolPath.CIRCLE,
        scale: 12,
        fillColor: DAY_COLORS[item._dayIdx % DAY_COLORS.length],
        fillOpacity: 1,
        strokeColor: '#fff',
        strokeWeight: 2
      }},
      label: {{
        text: item._letter || String(item.seq),
        color: 'white',
        fontWeight: '700',
        fontSize: '11px'
      }}
    }});

    // Day label on first point of each day (proper OverlayView)
    if (item._dayFirst) {{
      (function(txt, clr, latLng) {{
        var overlay = new google.maps.OverlayView();
        overlay.onAdd = function() {{
          var div = document.createElement('div');
          div.style.cssText = 'position:absolute;background:' + clr + ';color:white;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;white-space:nowrap;box-shadow:0 1px 4px rgba(0,0,0,0.3);pointer-events:none;';
          div.textContent = txt;
          this._div = div;
          this.getPanes().floatPane.appendChild(div);
        }};
        overlay.draw = function() {{
          var proj = this.getProjection();
          if (!proj || !this._div) return;
          var pixel = proj.fromLatLngToDivPixel(latLng);
          this._div.style.left = (pixel.x + 14) + 'px';
          this._div.style.top = (pixel.y - 28) + 'px';
        }};
        overlay.onRemove = function() {{
          if (this._div) {{
            this._div.parentNode.removeChild(this._div);
            this._div = null;
          }}
        }};
        overlay.setMap(mapInstance);
        currentOverlays.push(overlay);
      }})('Day ' + (item._dayIdx + 1), DAY_COLORS[item._dayIdx % DAY_COLORS.length], new google.maps.LatLng(lat, lng));
    }}

    // InfoWindow content
    var thumbHtml = '';
    if (item.img1) {{
      thumbHtml += '<img class="info-thumb" src="' + item.img1 + '" onerror="this.style.display=\\'none\\'" onclick="openLightbox(this.src)" style="width:80px;height:54px;object-fit:cover;border-radius:6px;cursor:pointer;margin-right:4px;">';
    }} else {{
      thumbHtml += '<div style="width:80px;height:54px;border-radius:6px;background:#f0f0f0;display:flex;align-items:center;justify-content:center;font-size:18px;color:#bbb;">📷</div>';
    }}
    if (item.img2) {{
      thumbHtml += '<img class="info-thumb" src="' + item.img2 + '" onerror="this.style.display=\\'none\\'" onclick="openLightbox(this.src)" style="width:80px;height:54px;object-fit:cover;border-radius:6px;cursor:pointer;">';
    }}
    var navLinks = '';
    var navPlatforms = {json.dumps(map_platforms)};
    navPlatforms.forEach(function(p) {{
      var uri = genNavUri(p, lat, lng, item.name);
      var names = {{'google':'Google','amap':'高德','baidu':'百度'}};
      var colors = {{'google':'#1a73e8','amap':'#008dfd','baidu':'#2932e1'}};
      navLinks += '<a href="' + uri + '" target="_blank" style="background:' + colors[p] + ';color:white;padding:2px 8px;border-radius:4px;font-size:12px;text-decoration:none;margin-right:4px;">' + names[p] + '</a>';
    }});

    var infoContent = '<div style="max-width:280px;">' +
      '<h4 style="margin:0 0 4px;font-size:14px;">' + (TYPE_ICONS[item.type] || '📍') + ' ' + item.name + '</h4>' +
      '<div style="font-size:12px;color:#666;margin-bottom:4px;">类型: ' + item.type + '</div>' +
      '<div style="display:flex;gap:4px;margin-bottom:6px;">' + thumbHtml + '</div>' +
      '<div style="display:flex;gap:4px;">' + navLinks + '</div>' +
      '</div>';

    var infoWindow = new google.maps.InfoWindow({{
      content: infoContent,
      maxWidth: 300
    }});

    marker.addListener('click', function() {{
      // Close all other info windows
      currentMarkers.forEach(function(m) {{
        if (m._infoWindow) m._infoWindow.close();
      }});
      infoWindow.open(mapInstance, marker);
    }});
    marker._infoWindow = infoWindow;

    bounds.extend({{lat: lat, lng: lng}});
    currentMarkers.push(marker);
    marker.setMap(mapInstance);
  }});

  // Draw polylines (Google Maps)
  days.forEach(function(day, dayIdx) {{
    var color = DAY_COLORS[dayIdx % DAY_COLORS.length];
    var items = routeData[day];
    var path = [];
    items.forEach(function(item) {{
      var parts = item.coords.split(',');
      path.push({{lat: parseFloat(parts[0]), lng: parseFloat(parts[1])}});
    }});
    if (path.length >= 2) {{
      var polyline = new google.maps.Polyline({{
        path: path,
        strokeColor: color,
        strokeWeight: 3,
        strokeOpacity: 0.8,
        geodesic: true
      }});
      currentPolylines.push(polyline);
      polyline.setMap(mapInstance);
    }}
  }});

  // Dashed lines between days (Google Maps)
  for (var i = 0; i < days.length - 1; i++) {{
    var dayItems1 = routeData[days[i]];
    var dayItems2 = routeData[days[i + 1]];
    var lastItem = dayItems1[dayItems1.length - 1];
    var firstItem = dayItems2[0];
    var p1 = lastItem.coords.split(',');
    var p2 = firstItem.coords.split(',');
    var dashLine = new google.maps.Polyline({{
      path: [
        {{lat: parseFloat(p1[0]), lng: parseFloat(p1[1])}},
        {{lat: parseFloat(p2[0]), lng: parseFloat(p2[1])}}
      ],
      strokeColor: '#999',
      strokeWeight: 2,
      strokeOpacity: 0.6,
      icons: [{{icon: {{path: 'M 0,-1 0,1', strokeOpacity: 1, strokeWeight: 2}}, offset: '0', repeat: '12px'}}]
    }});
    currentPolylines.push(dashLine);
    dashLine.setMap(mapInstance);
  }}

  if (currentMarkers.length > 0) {{
    mapInstance.fitBounds(bounds, {{top: 60, right: 60, bottom: 60, left: 60}});
  }}
}}

function flyTo(idx) {{
  document.querySelectorAll('.itin-item').forEach(function(el) {{ el.classList.remove('active'); }});
  var items = document.querySelectorAll('.itin-item');
  if (items[idx]) {{
    items[idx].classList.add('active');
  }}
  if (currentMarkers[idx]) {{
    if (mapInstance._engine === 'google') {{
      var pos = currentMarkers[idx].getPosition();
      mapInstance.setCenter(pos);
      mapInstance.setZoom(12);
      // Close all info windows then open this one
      currentMarkers.forEach(function(m) {{
        if (m._infoWindow) m._infoWindow.close();
      }});
      if (currentMarkers[idx]._infoWindow) {{
        currentMarkers[idx]._infoWindow.open(mapInstance, currentMarkers[idx]);
      }}
    }} else {{
      mapInstance.setZoomAndCenter(12, currentMarkers[idx].getPosition());
      currentMarkers[idx].infoWindow.open(mapInstance, currentMarkers[idx].getPosition());
    }}
  }}
}}

function typeIconHtml(item) {{
  return '<span style="color:' + (TYPE_COLORS[item.type] || '#9E9E9E') + '">' + (TYPE_ICONS[item.type] || '📍') + '</span>';
}}

function genNavUri(platform, lat, lng, name) {{
  if (platform === 'google') {{
    return 'https://www.google.com/maps/search/?api=1&query=' + lat + ',' + lng;
  }} else if (platform === 'amap') {{
    var gcj = wgs84ToGcj02(lat, lng);
    return 'https://uri.amap.com/marker?position=' + gcj.lng + ',' + gcj.lat + '&name=' + encodeURIComponent(name);
  }} else if (platform === 'baidu') {{
    var bd = wgs84ToBd09(lat, lng);
    return 'https://api.map.baidu.com/marker?location=' + bd.lat + ',' + bd.lng + '&title=' + encodeURIComponent(name) + '&content=&output=html&src=webapp.travelGuide';
  }}
  return '';
}}

// === Coordinate conversion (WGS84 -> GCJ02 -> BD09) ===
function wgs84ToGcj02(lat, lng) {{
  var a = 6378245.0;
  var ee = 0.00669342162296594323;
  var dLat = transformLat(lng - 105.0, lat - 35.0);
  var dLng = transformLng(lng - 105.0, lat - 35.0);
  var radLat = lat / 180.0 * Math.PI;
  var magic = Math.sin(radLat);
  magic = 1 - ee * magic * magic;
  var sqrtMagic = Math.sqrt(magic);
  dLat = (dLat * 180.0) / ((a * (1 - ee)) / (magic * sqrtMagic) * Math.PI);
  dLng = (dLng * 180.0) / (a / sqrtMagic * Math.cos(radLat) * Math.PI);
  return {{lat: lat + dLat, lng: lng + dLng}};
}}

function wgs84ToBd09(lat, lng) {{
  var gcj = wgs84ToGcj02(lat, lng);
  var x = gcj.lng, y = gcj.lat;
  var z = Math.sqrt(x * x + y * y) + 0.00002 * Math.sin(y * Math.PI * 3000.0 / 180.0);
  var theta = Math.atan2(y, x) + 0.000003 * Math.cos(x * Math.PI * 3000.0 / 180.0);
  return {{lat: z * Math.sin(theta) + 0.006, lng: z * Math.cos(theta) + 0.0065}};
}}

function transformLat(lng, lat) {{
  var ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + 0.1 * lng * lat + 0.2 * Math.sqrt(Math.abs(lng));
  ret += (20.0 * Math.sin(6.0 * lng * Math.PI) + 20.0 * Math.sin(2.0 * lng * Math.PI)) * 2.0 / 3.0;
  ret += (20.0 * Math.sin(lat * Math.PI) + 40.0 * Math.sin(lat / 3.0 * Math.PI)) * 2.0 / 3.0;
  ret += (160.0 * Math.sin(lat / 12.0 * Math.PI) + 320 * Math.sin(lat * Math.PI / 30.0)) * 2.0 / 3.0;
  return ret;
}}

function transformLng(lng, lat) {{
  var ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + 0.1 * lng * lat + 0.1 * Math.sqrt(Math.abs(lng));
  ret += (20.0 * Math.sin(6.0 * lng * Math.PI) + 20.0 * Math.sin(2.0 * lng * Math.PI)) * 2.0 / 3.0;
  ret += (20.0 * Math.sin(lng * Math.PI) + 40.0 * Math.sin(lng / 3.0 * Math.PI)) * 2.0 / 3.0;
  ret += (150.0 * Math.sin(lng / 12.0 * Math.PI) + 300.0 * Math.sin(lng / 30.0 * Math.PI)) * 2.0 / 3.0;
  return ret;
}}
</script>
</body>
</html>
""")

    html_text = '\n'.join(parts)
    if DEMO_MODE:
        demo_banner = (
            "Demo fixture — offline sample data; not live platform data"
            if OUTPUT_LANG == "en"
            else "Demo fixture — 离线演示数据；不是实时平台数据"
        )
        html_text = html_text.replace(
            "<header>",
            f'<header>\n<div style="background:#fff3cd;color:#856404;padding:8px 12px;border-radius:8px;margin-bottom:12px;font-size:13px">{demo_banner}</div>',
            1,
        )
    if OUTPUT_LANG == "en":
        # Keep source-derived place descriptions untouched, but localize the
        # stable UI labels so the English artifact is usable rather than a
        # machine-translated copy of the data.
        html_translations = {
            "lang=\"zh-CN\"": "lang=\"en\"", "旅行攻略": "Travel Guide",
            "卡片列表": "Cards", "地图路线": "Map Route", "展开/收起全部": "Expand/Collapse All",
            "搜索目的地/酒店/餐厅/景点...": "Search destinations, hotels, restaurants, attractions...",
            "点击地点右侧": "Use the map buttons to navigate; click destination or category headers to expand/collapse",
            "按钮可一键导航；点击目的地标题可展开/收起": "",
            ">住宿<": ">Accommodation<", ">美食<": ">Food<", ">景点<": ">Attractions<", ">潜水<": ">Diving<", ">交通<": ">Transport<",
            "🏨 住宿 ": "🏨 Accommodation ", "🍽 美食 ": "🍽 Food ", "📍 景点 ": "📍 Attractions ", "🤿 潜水 ": "🤿 Diving ", "🚌 交通 ": "🚌 Transport ",
            "目的地</div>": "Destinations</div>", "目的地总览": "Destinations", "自然/文化体验": "Nature / Culture Experience",
            "Google评分：": "Google rating: ", "搜索补充：": "Search supplement: ", "条": " reviews",
            "互动探索": "Interactive guide", "一键导航": "navigation", "按目的地分类": "by destination",
            "元/晚": " CNY/night", "人均": "Avg. CNY ", "元": " CNY",
            "个地点": " places", "个景点": " attractions", "家酒店": " hotels", "家餐厅": " restaurants",
            "行程路线": "Itinerary", "方案": "Option", "数据仅供参考": "Data for reference only",
            "导航平台": "Navigation platforms", "高德地图": "Amap", "百度地图": "Baidu Maps", "Google Maps": "Google Maps",
            "高德评分": "Amap rating", "搜索补充": "Search supplement", "AI推荐分": "AI recommendation",
            "AI提供，仅供参考；非平台评分": "AI-provided, for reference only; not a platform rating",
            "需要高德JS API Key": "Amap JS API key required", "海外目的地需提供Google Maps API Key": "Google Maps API key required for overseas destinations",
            "需Key": "key required",
            "该路线暂无详细行程编排": "This route has no detailed itinerary yet",
            "仅推荐路线提供每日行程地图展示": "Only the recommended route has a daily map itinerary",
        }
        for source, target in html_translations.items():
            html_text = html_text.replace(source, target)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_text)
    print(f"HTML saved: {output_path}")
    if map_route_enabled:
        engine = 'Google Maps JS API' if overseas else 'AMap JS API'
        print(f"  Map route view: ENABLED ({engine})")
    elif has_itinerary and not amap_js_key and not (overseas and google_maps_key):
        print(f"  Map route view: DISABLED (no --amap-js-key)")
    elif has_itinerary and overseas and not google_maps_key:
        print(f"  Map route view: DISABLED (overseas, need --google-maps-key)")
    else:
        print(f"  Map route view: DISABLED (no ITINERARY data)")


# ============================================================
# MARKDOWN GENERATION
# ============================================================

def generate_markdown(output_path):
    """Generate comprehensive Markdown travel guide."""
    guide_title = f"{country_display()} Travel Guide" if OUTPUT_LANG == "en" else f"{COUNTRY}旅行攻略大全"
    md = []
    md.append(f"# {guide_title}")
    if DEMO_MODE:
        md.append("> **Demo fixture / 离线演示数据：** This guide is generated from sanitized sample data and is not live platform data.")
    md.append("")
    md.append("---")
    md.append("")

    # 1. 目的地总览
    md.append("## 1. 目的地总览")
    md.append("")
    md.append(f"共 **{len(DESTINATIONS)}** 个目的地：")
    md.append("")
    md.append("| 中文名 | 英文名 | 区域 | 核心体验 | 建议天数 | 最佳季节 | 到达方式 | 是否纳入环线 |")
    md.append("|---|---|---|---|---|---|---|---|")
    for d in DESTINATIONS:
        md.append(f"| {d[0]} | {d[1]} | {d[2]} | {d[3]} | {d[4]} | {d[5]} | {d[6]} | {d[7]} |")
    md.append("")
    md.append("---")
    md.append("")

    # 2. 路线方案
    md.append("## 2. 路线方案")
    md.append("")
    md.append("| 方案名称 | 天数 | 路线 | 特色 | 适合人群 | 预算(元) | 交通方式 |")
    md.append("|---|---|---|---|---|---|---|")
    for r in ROUTES:
        md.append(f"| {r[0]} | {r[1]} | {r[2]} | {r[3]} | {r[4]} | {r[5]} | {r[6]} |")
    md.append("")
    md.append("---")
    md.append("")

    # 每日行程详情 (新增)
    if ITINERARY:
        md.append("## 每日行程详情")
        md.append("")
        md.append("> 仅展示推荐路线的每日行程点位顺序")
        md.append("")
        routes_itin = {}
        for item in ITINERARY:
            r_name = item[0]
            if r_name not in routes_itin:
                routes_itin[r_name] = []
            routes_itin[r_name].append(item)
        for r_name, items in routes_itin.items():
            md.append(f"### {r_name}")
            md.append("")
            md.append("| 日序 | 序号 | 类型 | 地点 | 经纬度 |")
            md.append("|---|---|---|---|---|")
            for item in items:
                md.append(f"| {item[1]} | {item[2]} | {item[3]} | {item[4]} | {item[5]} |")
            md.append("")
        md.append("---")
        md.append("")

    # 3. 住宿推荐
    md.append("## 3. 住宿推荐")
    md.append("")
    md.append(f"共 **{len(HOTELS)}** 家酒店：")
    md.append("")
    for dest_name in [d[0] for d in DESTINATIONS]:
        dest_hotels = [h for h in HOTELS if h[0] == dest_name]
        if not dest_hotels:
            continue
        md.append(f"### {destination_display(dest_name)}")
        md.append("")
        md.append("| 酒店名称 | 类型 | 参考价(元/晚) | 评分 | 评价数 | 特色 |")
        md.append("|---|---|---|---|---|---|")
        for h in dest_hotels:
            md.append(f"| {display_name(h[1], h[2])} | {h[3]} | {h[4]} | ⭐{h[5]} | {h[6]} | {h[10]} |")
        md.append("")
    md.append("---")
    md.append("")

    # 4. 美食推荐
    md.append("## 4. 美食推荐")
    md.append("")
    md.append(f"共 **{len(RESTAURANTS)}** 家餐厅：")
    md.append("")
    for dest_name in [d[0] for d in DESTINATIONS]:
        dest_rest = [r for r in RESTAURANTS if r[0] == dest_name]
        if not dest_rest:
            continue
        md.append(f"### {destination_display(dest_name)}")
        md.append("")
        md.append("| 餐厅名称 | 特色菜品 | 人均(元) | 评分 | 位置 |")
        md.append("|---|---|---|---|---|")
        for r in dest_rest:
            md.append(f"| {display_name(r[1], r[2])} | {r[3]} | {r[4]} | ⭐{r[5]} | {r[10]} |")
        md.append("")
    md.append("---")
    md.append("")

    # 5. 交通信息
    md.append("## 5. 交通信息")
    md.append("")
    md.append("| 路线 | 交通方式 | 运营公司 | 参考票价(元) | 行程时间 | 班次频率 | 备注 |")
    md.append("|---|---|---|---|---|---|---|")
    for t in TRANSPORT:
        md.append(f"| {t[0]} | {t[1]} | {t[2]} | {t[3]} | {t[4]} | {t[5]} | {t[8]} |")
    md.append("")
    md.append("---")
    md.append("")

    # 6. 景点活动
    md.append("## 6. 景点活动")
    md.append("")
    md.append(f"共 **{len(ATTRACTIONS)}** 个景点/活动：")
    md.append("")
    for dest_name in [d[0] for d in DESTINATIONS]:
        dest_attr = [a for a in ATTRACTIONS if a[0] == dest_name]
        if not dest_attr:
            continue
        md.append(f"### {destination_display(dest_name)}")
        md.append("")
        md.append("| 景点名称 | 类型 | 门票(元) | 评分 | 游玩时间 | 描述 |")
        md.append("|---|---|---|---|---|---|")
        for a in dest_attr:
            md.append(f"| {display_name(a[1], a[2])} | {a[3]} | {a[4]} | ⭐{a[5]} | {a[9]} | {a[10]} |")
        md.append("")
    md.append("---")
    md.append("")

    # 7. 潜水活动
    if DIVE_SITES:
        md.append("## 7. 潜水活动")
        md.append("")
        md.append("| 目的地 | 潜点名称 | 类型 | 价格(元) | 评分 | 适合等级 | 描述 |")
        md.append("|---|---|---|---|---|---|---|")
        for d in DIVE_SITES:
            md.append(f"| {d[0]} | {d[1]}({d[2]}) | {d[3]} | {d[4]} | ⭐{d[5]} | {d[9]} | {d[10]} |")
        md.append("")
        md.append("---")
        md.append("")

    # 8. 预算汇总
    md.append("## 8. 预算汇总")
    md.append("")
    md.append("| 天数 | 日期 | 行程 | 住宿 | 餐饮 | 交通 | 活动 | 其他 | 日合计 |")
    md.append("|---|---|---|---|---|---|---|---|---|")
    for b in BUDGET:
        md.append(f"| {b[0]} | {b[1]} | {b[2]} | {b[3]} | {b[4]} | {b[5]} | {b[6]} | {b[7]} | {b[8]} |")
    md.append("")
    md.append("---")
    md.append("")

    # 9. 实用信息
    md.append("## 9. 实用信息")
    md.append("")
    for p in PRACTICAL:
        md.append(f"- **{p[0]}：** {p[1]}")
    md.append("")
    md.append("---")
    md.append("")
    md.append(f"*{guide_title}*")

    md_text = '\n'.join(md)
    if OUTPUT_LANG == "en":
        md_translations = {
            "旅行攻略大全": "Travel Guide",
            "目的地总览": "Destinations", "共 **": "Total: **", "个目的地：": " destinations:",
            "路线方案": "Route Options", "每日行程详情": "Daily Itinerary", "仅展示推荐路线的每日行程点位顺序": "Daily itinerary for the recommended route",
            "住宿推荐": "Accommodation", "共 **": "Total: **", "家酒店：": " hotels:",
            "美食推荐": "Food & Restaurants", "家餐厅：": " restaurants:", "交通信息": "Transport",
            "景点活动": "Attractions & Activities", "个景点/活动：": " attractions/activities:",
            "潜水活动": "Diving", "预算汇总": "Budget Summary", "实用信息": "Practical Information",
            "中文名": "Chinese Name", "英文名": "English Name", "区域": "Region", "核心体验": "Key Experiences",
            "建议天数": "Suggested Days", "最佳季节": "Best Season", "到达方式": "Arrival", "是否纳入环线": "Included in Route",
        }
        md_translations.update({
            "| 方案名称 | 天数 | 路线 | 特色 | 适合人群 | 预算(元) | 交通方式 |": "| Route Option | Days | Route | Highlights | Best For | Budget (CNY) | Transport |",
            "| 酒店名称 | 类型 | 参考价(元/晚) | 评分 | 评价数 | 特色 |": "| Hotel | Type | Reference Price (CNY/night) | Rating | Reviews | Highlights |",
            "| 餐厅名称 | 特色菜品 | 人均(元) | 评分 | 位置 |": "| Restaurant | Signature Dishes | Average Spend (CNY) | Rating | Location |",
            "| 景点名称 | 类型 | 门票(元) | 评分 | 游玩时间 | 描述 |": "| Attraction | Type | Ticket (CNY) | Rating | Suggested Time | Description |",
            "| 路线 | 交通方式 | 运营公司 | 参考票价(元) | 行程时间 | 班次频率 | 备注 |": "| Route | Transport | Operator | Reference Fare (CNY) | Travel Time | Frequency | Notes |",
            "| 天数 | 日期 | 行程 | 住宿 | 餐饮 | 交通 | 活动 | 其他 | 日合计 |": "| Days | Date | Plan | Lodging | Food | Transport | Activities | Other | Daily Total |",
        })
        for source, target in md_translations.items():
            md_text = md_text.replace(source, target)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(md_text)
    print(f"Markdown saved: {output_path}")


# ============================================================
# MAIN
# ============================================================

def validate_data_integrity(expected_days=None, require_images=True):
    """Fail before delivery when the data model would create a broken guide."""
    _require_formal_run_manifest()
    errors = []
    allowed_categories = {"景点", "餐厅", "酒店", "潜水", "交通"}
    for i, row in enumerate(TRANSPORT, 1):
        if len(row) < 9:
            errors.append(f"TRANSPORT[{i}] 字段数量不足，必须包含9个字段")
            continue
        if any(not str(row[idx]).strip() for idx in range(6)):
            errors.append(f"TRANSPORT[{i}] 缺少路线/方式/运营方/票价/耗时/班次字段")
    poi_rows = [("HOTELS", HOTELS, 8, 9, 4), ("RESTAURANTS", RESTAURANTS, 8, 9, 4), ("ATTRACTIONS", ATTRACTIONS, 7, 8, 4)]
    poi_names = set()
    for label, rows, img1_idx, img2_idx, price_idx in poi_rows:
        for i, row in enumerate(rows, 1):
            if len(row) <= max(img2_idx, price_idx):
                errors.append(f"{label}[{i}] 字段数量不足")
                continue
            name = str(row[1]).strip() if len(row) > 1 else ""
            if not name:
                errors.append(f"{label}[{i}] 点位名称为空")
            poi_names.add(name)
            price = str(row[price_idx]).strip()
            if not price or price in {"待核验", "待补充", "未知", "N/A"}:
                errors.append(f"{label}[{i}] {name or '未命名'} 价格仍是占位值")
            if require_images and (not str(row[img1_idx]).startswith(("http://", "https://")) or not str(row[img2_idx]).startswith(("http://", "https://"))):
                errors.append(f"{label}[{i}] {name or '未命名'} 缺少两张可访问图片URL")

    route_names = [str(r[0]).strip() for r in ROUTES if r and r[0]]
    itinerary_routes = {}
    for i, item in enumerate(ITINERARY, 1):
        if len(item) < 6:
            errors.append(f"ITINERARY[{i}] 字段数量不足")
            continue
        route, day, _, category, name, coords = item[:6]
        itinerary_routes.setdefault(str(route), set()).add(str(day))
        if str(category).strip() not in allowed_categories:
            errors.append(f"ITINERARY[{i}] 未知点位类型: {category}")
        if str(category).strip() == "交通":
            transport_names = {str(row[0]).strip() for row in TRANSPORT if row and row[0]}
            if str(name).strip() not in transport_names:
                errors.append(f"ITINERARY[{i}] 交通点 {name} 不存在于 TRANSPORT 数据")
        elif str(name).strip() not in poi_names:
            errors.append(f"ITINERARY[{i}] 点位 {name} 不存在于 POI 数据")
        if not parse_coords(str(coords))[0] is not None:
            errors.append(f"ITINERARY[{i}] 坐标格式错误: {coords}")

    expected = expected_days
    if expected is None and itinerary_routes:
        expected = max((int(d.replace("Day ", "")) for days in itinerary_routes.values() for d in days if str(d).startswith("Day ")), default=0)
    for route in route_names:
        days = itinerary_routes.get(route, set())
        if not days:
            errors.append(f"路线 {route} 没有每日行程数据")
            continue
        if expected:
            actual = {int(str(d).replace("Day ", "")) for d in days if str(d).startswith("Day ")}
            missing = set(range(1, expected + 1)) - actual
            if missing:
                errors.append(f"路线 {route} 缺少天数: {sorted(missing)}")

    global TRAVEL_MODE
    if TRAVEL_MODE not in {"自驾", "非自驾"}:
        TRAVEL_MODE = "非自驾"
    require_itinerary_transport = not (DOMESTIC_DESTINATION and effective_travel_mode() == "自驾")
    # TRANSPORT is not merely a separate worksheet. Every collected inter-city
    # edge must be visible in the daily itinerary so the map and route panel
    # explain how the traveller actually moves between destinations.
    transport_names = {str(row[0]).strip() for row in TRANSPORT if row and row[0]}
    itinerary_transport_names = {
        str(item[4]).strip() for item in ITINERARY
        if len(item) >= 6 and str(item[3]).strip() == "交通"
    }
    missing_transport = sorted(transport_names - itinerary_transport_names)
    if require_itinerary_transport and missing_transport:
        errors.append(f"TRANSPORT 中的交通边未进入 ITINERARY: {missing_transport}")
    if require_itinerary_transport and transport_names:
        for route in route_names:
            route_transport_count = sum(
                1 for item in ITINERARY
                if len(item) >= 6 and str(item[0]).strip() == route and str(item[3]).strip() == "交通"
            )
            if route_transport_count == 0:
                errors.append(f"路线 {route} 含跨目的地规划但没有每日交通节点")

    if errors:
        raise ValueError("数据完整性校验失败:\n- " + "\n- ".join(errors[:40]) + (f"\n- ...另有 {len(errors)-40} 项" if len(errors) > 40 else ""))
    return True

def _require_formal_run_manifest():
    """Prevent ad-hoc builders from producing delivery artifacts."""
    if os.environ.get("TRAVEL_SCOPE_FORMAL_RUN") != "1":
        raise ValueError("正式生成已阻断：缺少 TRAVEL_SCOPE_FORMAL_RUN=1；临时生成器不得直接产出攻略")
    manifest_path = os.environ.get("TRAVEL_SCOPE_RUN_MANIFEST", "").strip()
    if not manifest_path or not os.path.isfile(manifest_path):
        raise ValueError("正式生成已阻断：缺少运行 manifest，必须先完成标准 SOP 阶段")
    try:
        with open(manifest_path, "r", encoding="utf-8") as fh:
            manifest = json.load(fh)
    except Exception as exc:
        raise ValueError(f"正式生成已阻断：运行 manifest 无法读取: {exc}") from exc
    required = {"destination_discovery", "route_exploration", "source_search", "source_audit", "qa"}
    stages = manifest.get("completed_stages") or []
    missing = sorted(required - set(stages))
    if missing:
        raise ValueError(f"正式生成已阻断：SOP 阶段未完成: {missing}")
    if manifest.get("source_audit_passed") is not True:
        raise ValueError("正式生成已阻断：数据源审计未通过")
    evidence_path = manifest.get("source_evidence_path")
    if not evidence_path or not os.path.isfile(evidence_path):
        raise ValueError("正式生成已阻断：缺少来源证据台账")

def main():
    parser = argparse.ArgumentParser(description='Generate travel guide Excel + HTML + Markdown')
    parser.add_argument('--output-dir', default='.', help='Output directory')
    parser.add_argument('--map-platform', default='all',
                        choices=['google', 'amap', 'baidu', 'all'],
                        help='Map platform for HTML navigation buttons')
    parser.add_argument('--country', default=None, help='Override country name for filenames')
    parser.add_argument('--amap-js-key', default='', help='AMap JS API key for map route view')
    parser.add_argument('--amap-security', default='', help='AMap JS API security code')
    parser.add_argument('--google-maps-key', default='', help='Google Maps JS API key for overseas map route view')
    parser.add_argument('--language', default='cn', choices=['cn', 'en', 'both'],
                        help='Output language: cn, en, or both (default: cn)')
    parser.add_argument('--mode', default='live', choices=['live', 'demo'],
                        help='live uses collected/provider data; demo uses offline fixture data')
    parser.add_argument('--fixture', default=None,
                        help='Demo fixture JSON path (defaults to fixtures/demo.json)')
    parser.add_argument('--expected-days', type=int, default=None, help='Expected trip days for strict itinerary validation')
    parser.add_argument('--no-strict-qa', action='store_true', help='Allow incomplete data; not recommended for delivery')
    args = parser.parse_args()

    global COUNTRY, OUTPUT_LANG, DEMO_MODE
    DEMO_MODE = args.mode == 'demo'
    if DEMO_MODE:
        fixture_path = args.fixture or os.path.join(os.path.dirname(__file__), '..', 'fixtures', 'demo.json')
        fixture_path = os.path.abspath(fixture_path)
        if not os.path.isfile(fixture_path):
            print(f"ERROR: Demo fixture not found: {fixture_path}")
            sys.exit(2)
        load_demo_fixture(fixture_path)
    if args.country:
        COUNTRY = args.country

    # Determine map platforms
    if args.map_platform == 'all':
        map_platforms = ['google', 'amap', 'baidu']
    else:
        map_platforms = [args.map_platform]

    output_dir = os.path.abspath(args.output_dir)
    os.makedirs(output_dir, exist_ok=True)

    # Validate data
    if not DESTINATIONS or len(DESTINATIONS) < 1:
        print("ERROR: DESTINATIONS array is empty or too small. LLM must fill in real data before running.")
        print("Edit the DATA section in this script with real destination data.")
        sys.exit(1)

    if DEMO_MODE:
        print("[DEMO] Offline fixture mode: no provider APIs or API keys are used.")
        print("[DEMO] Formal production source-audit gate is not applicable to fixture output.")
    elif not args.no_strict_qa:
        try:
            validate_data_integrity(expected_days=args.expected_days)
        except ValueError as exc:
            print(f"ERROR: {exc}")
            print("Use --no-strict-qa only for an explicitly labelled research draft.")
            sys.exit(2)

    # Generate one or both language artifacts from the same validated data model.
    print(f"Generating travel guide for: {COUNTRY}")
    print(f"Map platforms: {', '.join(map_platforms)}")
    print(f"Output directory: {output_dir}")
    print()

    languages = ['cn', 'en'] if args.language == 'both' else [args.language]
    generated = []
    for lang in languages:
        OUTPUT_LANG = lang
        suffix = "" if lang == "cn" else "-en"
        excel_path = os.path.join(output_dir, f"{COUNTRY}旅行攻略大全{suffix}.xlsx")
        research_excel_path = os.path.join(output_dir, f"{COUNTRY}旅行攻略研究版{suffix}.xlsx")
        html_path = os.path.join(output_dir, f"{COUNTRY}旅行攻略-地图标记{suffix}.html")
        md_path = os.path.join(output_dir, f"{COUNTRY}旅行攻略大全{suffix}.md")
        generate_excel(excel_path, include_research=False)
        generate_excel(research_excel_path, include_research=True)
        print()
        generate_html(html_path, map_platforms, args.amap_js_key, args.amap_security, args.google_maps_key)
        print()
        generate_markdown(md_path)
        generated.extend([excel_path, research_excel_path, html_path, md_path])

    print(f"\n=== SUMMARY ===")
    print(f"Country: {COUNTRY}")
    print(f"Mode: {'demo fixture' if DEMO_MODE else 'live/provider data'}")
    print(f"Destinations: {len(DESTINATIONS)}")
    print(f"Hotels: {len(HOTELS)}")
    print(f"Restaurants: {len(RESTAURANTS)}")
    print(f"Attractions: {len(ATTRACTIONS)}")
    print(f"Dive Sites: {len(DIVE_SITES)}")
    print(f"Transport: {len(TRANSPORT)}")
    print(f"Practical Info: {len(PRACTICAL)}")
    print(f"Itinerary Items: {len(ITINERARY)}")
    print(f"Map platforms: {', '.join(map_platforms)}")
    print(f"\nFiles generated:")
    for index, path in enumerate(generated, 1):
        print(f"  {index}. {path}")


if __name__ == '__main__':
    main()
