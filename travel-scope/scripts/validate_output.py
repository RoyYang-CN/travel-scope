#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Static delivery gate for Travel-Scope HTML/Excel outputs."""

import argparse
import json
import re
import sys
from pathlib import Path


def fail(errors, message):
    errors.append(message)


def main():
    parser = argparse.ArgumentParser(description="Validate Travel-Scope delivery artifacts")
    parser.add_argument("--html", required=True)
    parser.add_argument("--delivery-xlsx")
    parser.add_argument("--research-xlsx")
    parser.add_argument("--scope-json")
    parser.add_argument("--route-plan-json", help="Optional route semantics/coverage plan")
    parser.add_argument("--expected-days", type=int)
    args = parser.parse_args()

    errors = []
    html_path = Path(args.html)
    if not html_path.exists():
        print(f"ERROR: HTML not found: {html_path}")
        return 2
    html = html_path.read_text(encoding="utf-8", errors="replace")

    card_blocks = html.split('<div class="loc-card"')[1:]
    cards = len(card_blocks)
    poi_cards = sum('class="loc-thumb"' in block for block in card_blocks)
    if not cards:
        fail(errors, "HTML 未生成任何卡片")
    image_tags = re.findall(r'<img class="loc-thumb"\s+src="([^"].*?)"', html)
    if poi_cards and len(image_tags) != poi_cards * 2:
        fail(errors, f"图片数量不完整: poi_cards={poi_cards}, image_tags={len(image_tags)}, expected={poi_cards * 2}")
    if any(not u.startswith(("http://", "https://")) for u in image_tags):
        fail(errors, "存在非 HTTP(S) 图片URL")
    detail_blocks = re.findall(r'<div class="loc-detail">(.*?)</div>', html, re.S)
    # 国内评分可能因来源访问限制而明确标记为“待核验”，这不等于价格占位。
    # 卡片结构为“类型 | 价格 | 评分”，门禁只阻断价格字段的占位值。
    price_fields = [b.split("|")[1].strip() for b in detail_blocks if len(b.split("|")) >= 2]
    if any("待核验" in b or "待补充" in b for b in price_fields):
        fail(errors, "卡片价格字段仍包含占位值")

    destination_names = set(re.findall(r'<div class="dest-section" data-dest="([^"]+)"', html))
    if args.scope_json:
        scope = json.loads(Path(args.scope_json).read_text(encoding="utf-8"))
        expected_destinations = {
            x.get("name") for x in scope.get("selected_destinations", [])
            if x.get("status") in {"正式纳入", "路线备选"}
        }
        missing = sorted(expected_destinations - destination_names)
        if missing:
            fail(errors, f"HTML遗漏已纳入/备选目的地: {missing}")

    route_match = re.search(r'var itineraryData = (\{.*?\});\s*\nvar allRouteNames = (\[.*?\]);', html, re.S)
    if not route_match:
        fail(errors, "未找到 itineraryData 或 allRouteNames")
    else:
        itinerary = json.loads(route_match.group(1))
        route_names = json.loads(route_match.group(2))
        expected_days = args.expected_days
        if expected_days is None:
            expected_days = max((int(d.replace("Day ", "")) for r in itinerary.values() for d in r), default=0)
        for route in route_names:
            if route not in itinerary:
                fail(errors, f"路线 {route} 没有 itineraryData")
                continue
            actual_days = {int(d.replace("Day ", "")) for d in itinerary[route] if d.startswith("Day ")}
            missing_days = set(range(1, expected_days + 1)) - actual_days
            if missing_days:
                fail(errors, f"路线 {route} 缺少天数: {sorted(missing_days)}")
            for day, points in itinerary[route].items():
                if len(points) < 2:
                    fail(errors, f"路线 {route} {day} 少于2个行程点")
                for point in points:
                    if str(point.get("type", point.get("category", ""))).strip() == "交通":
                        continue
                    img1 = str(point.get("img1") or "")
                    img2 = str(point.get("img2") or "")
                    if not img1.startswith(("http://", "https://")) or not img2.startswith(("http://", "https://")):
                        fail(errors, f"路线 {route} {day} 点位 {point.get('name', '')} 缺少双图")

        # A transport card/table alone is insufficient: when the delivery
        # contains transport data, every route must expose at least one
        # transport item in its daily route panel.
        transport_cards = bool(re.search(r'data-cat="交通"', html))
        route_transport_counts = {
            route: sum(
                1 for points in route_days.values() for point in points
                if str(point.get("type", point.get("category", ""))).strip() == "交通"
            )
            for route, route_days in itinerary.items()
        }
        if transport_cards:
            if not any(route_transport_counts.values()):
                fail(errors, "HTML 存在交通数据，但 ITINERARY 没有交通节点")
            for route, count in route_transport_counts.items():
                if count == 0:
                    fail(errors, f"路线 {route} 存在交通数据但每日行程没有交通节点")

        # A complete calendar is not enough: reject synthetic routes that
        # repeat the same two POIs for the entire trip or duplicate another route.
        route_signatures = {}
        min_unique = max(4, min(8, (expected_days + 1) // 2))
        for route in route_names:
            if route not in itinerary:
                continue
            day_values = list(itinerary[route].values())
            names = [str(p.get("name", "")).strip() for pts in day_values for p in pts if str(p.get("name", "")).strip()]
            unique_names = set(names)
            if len(unique_names) < min_unique:
                fail(errors, f"路线 {route} 点位过度重复: unique_points={len(unique_names)}, minimum={min_unique}")
            day_signatures = {tuple(sorted(str(p.get("name", "")).strip() for p in pts)) for pts in day_values}
            if len(day_signatures) < max(3, min(6, expected_days // 3)):
                fail(errors, f"路线 {route} 每日结构过度重复: unique_day_signatures={len(day_signatures)}")
            route_signatures[route] = unique_names
            for name in unique_names:
                day_count = sum(1 for pts in day_values if any(str(p.get("name", "")).strip() == name for p in pts))
                if day_count > max(3, expected_days // 2):
                    fail(errors, f"路线 {route} 点位 {name} 占用过多天数: {day_count}/{expected_days}")
        route_list = list(route_signatures)
        for i, left in enumerate(route_list):
            for right in route_list[i + 1:]:
                union = route_signatures[left] | route_signatures[right]
                overlap = len(route_signatures[left] & route_signatures[right]) / len(union) if union else 1.0
                if overlap >= 0.95:
                    fail(errors, f"路线 {left} 与 {right} 实际点位高度重复: overlap={overlap:.2f}")

        if args.route_plan_json:
            plan = json.loads(Path(args.route_plan_json).read_text(encoding="utf-8"))
            for route_plan in plan.get("routes", []):
                route = route_plan.get("name")
                if route not in itinerary:
                    fail(errors, f"路线计划 {route} 未出现在 HTML itineraryData")
                    continue
                route_text = " ".join(str(p.get("name", "")) for pts in itinerary[route].values() for p in pts)
                missing_required = [x for x in route_plan.get("required_poi_names", []) if x not in route_text]
                if missing_required:
                    fail(errors, f"路线 {route} 缺少路线计划要求的经典节点: {missing_required}")

    if args.delivery_xlsx or args.research_xlsx:
        try:
            import openpyxl
            if args.delivery_xlsx:
                wb = openpyxl.load_workbook(args.delivery_xlsx, read_only=True)
                if len(wb.sheetnames) != 11:
                    fail(errors, f"交付版工作表数量错误: {len(wb.sheetnames)}")
            if args.research_xlsx:
                wb = openpyxl.load_workbook(args.research_xlsx, read_only=True)
                if len(wb.sheetnames) != 14:
                    fail(errors, f"研究版工作表数量错误: {len(wb.sheetnames)}")
        except ImportError:
            fail(errors, "缺少 openpyxl，无法验证 Excel")

    if errors:
        print("FAIL")
        for e in errors:
            print(f"- {e}")
        return 1
    print(json.dumps({"status": "PASS", "destinations": len(destination_names), "cards": cards, "images": len(image_tags), "routes": len(route_names) if route_match else 0}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
